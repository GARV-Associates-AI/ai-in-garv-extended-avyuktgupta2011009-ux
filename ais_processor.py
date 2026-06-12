"""
ais_processor.py — Day 16 (Phase 4: Master DB + Fast Cache)
AIS (Annual Information Statement) PDF Reader & Reconciler

Matching strategy:
  Tier 1: ISIN match           → AUTO ACCEPT
  Tier 2: Exact normalized name → AUTO ACCEPT
  Tier 2.5: Master DB match     → AUTO ACCEPT (fast cache)
  Tier 3: Candidate matching (Acronym + Substring + Fuzzy 70%+):
            - 1 candidate  → AUTO ACCEPT
            - 2+ candidates → ASK USER (ambiguous)
"""

import re
from difflib import SequenceMatcher
import pypdf

# Day 16 — Fast master DB cache
from database import build_master_db_cache, resolve_isin_from_cache


# ============================================================
# NOISE WORDS
# ============================================================

NOISE_WORDS = {
    'LTD', 'LIMITED', 'PVT', 'PRIVATE', 'CO', 'COMPANY',
    'CORP', 'CORPORATION', 'INC', 'THE', 'AND', '&',
    'EQ', 'NEW', 'FV', 'RS', 'INDIA', 'INDIAN',
    'GROUP', 'HOLDINGS', 'INDUSTRIES', 'ENTERPRISES',
    'INTERNATIONAL', 'GLOBAL', '10/-', '2/-', '5/-', '1/-',
    'OF', 'IN', 'FOR', 'F.V.RE.1'
}


# ============================================================
# STEP 1 — EXTRACT PDF TEXT
# ============================================================

def extract_pdf_text(pdf_file):
    try:
        reader = pypdf.PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"AIS PDF read error: {e}")
        return None


# ============================================================
# STEP 2 — PARSE AIS SECTIONS
# ============================================================

def parse_ais_pdf(text):
    result = {'sells': [], 'buys': []}
    if not text:
        return result

    sales_text = extract_sales_section(text)
    if sales_text:
        result['sells'] = parse_sales_rows(sales_text)

    purchase_text = extract_purchase_section(text)
    if purchase_text:
        result['buys'] = parse_sales_rows(purchase_text)

    return result


def extract_sales_section(text):
    start_markers = [
        'Sale of securities and units of mutual fund',
        'Sale of listed equity',
    ]
    end_markers = [
        'Purchase of securities and units of mutual funds',
        'Purchase of securities',
        'Part B7', 'Part B3', 'Part B4',
    ]

    start_pos = -1
    for m in start_markers:
        p = text.find(m)
        if p >= 0:
            start_pos = p
            break
    if start_pos < 0:
        return None

    end_pos = len(text)
    for m in end_markers:
        p = text.find(m, start_pos + 50)
        if 0 <= p < end_pos:
            end_pos = p

    return text[start_pos:end_pos]


def extract_purchase_section(text):
    start_markers = [
        'Purchase of securities and units of mutual funds',
        'Purchase of securities',
    ]
    end_markers = ['Part B7', 'Part B3', 'Part B4', 'Refund']

    start_pos = -1
    for m in start_markers:
        p = text.find(m)
        if p >= 0:
            start_pos = p
            break
    if start_pos < 0:
        return None

    end_pos = len(text)
    for m in end_markers:
        p = text.find(m, start_pos + 50)
        if 0 <= p < end_pos:
            end_pos = p

    return text[start_pos:end_pos]


def parse_sales_rows(section_text):
    transactions = []
    if not section_text:
        return transactions

    flat_text = re.sub(r'\s+', ' ', section_text)
    isin_pattern = re.compile(r'\(IN[EFA90][A-Z0-9]{9}\)')
    date_pattern = re.compile(r'\b(\d{2}/\d{2}/\d{4})\b')

    isin_matches = list(isin_pattern.finditer(flat_text))
    if not isin_matches:
        return transactions

    for i, isin_match in enumerate(isin_matches):
        isin_start = isin_match.start()
        isin_end = isin_match.end()
        isin = isin_match.group(0).strip('()')

        if i == 0:
            row_start = 0
        else:
            row_start = isin_matches[i-1].end()

        if i == len(isin_matches) - 1:
            row_end = len(flat_text)
        else:
            next_isin_start = isin_matches[i+1].start()
            between = flat_text[isin_end:next_isin_start]
            dates_in_between = list(date_pattern.finditer(between))
            if dates_in_between:
                last_date = dates_in_between[-1]
                row_end = isin_end + last_date.start()
            else:
                row_end = next_isin_start

        before_isin = flat_text[row_start:isin_start]
        dates_in_row = list(date_pattern.finditer(before_isin))
        if not dates_in_row:
            continue
        date_match = dates_in_row[-1]
        date_str = date_match.group(1)

        company_raw = before_isin[date_match.end():].strip()
        company = clean_company_name(company_raw)

        if not company or len(company) < 2:
            continue

        after_isin = flat_text[isin_end:row_end]

        num_pattern = re.compile(r'[\d,]+\.\d+|\d{1,3}(?:,\d{3})+|\d+')
        nums = num_pattern.findall(after_isin)
        clean_nums = []
        for n in nums:
            try:
                val = float(n.replace(',', ''))
                clean_nums.append(val)
            except ValueError:
                pass

        if len(clean_nums) < 3:
            continue

        quantity = clean_nums[0]
        value = clean_nums[2] if len(clean_nums) >= 3 else clean_nums[1]

        if quantity <= 0 or value <= 0:
            continue

        transactions.append({
            'date': normalize_date(date_str),
            'company': company,
            'isin': isin,
            'quantity': quantity,
            'value': value,
            'type': 'SELL'
        })

    return transactions


def normalize_date(date_str):
    try:
        parts = date_str.split('/')
        if len(parts) == 3:
            d, m, y = parts
            return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    except Exception:
        pass
    return date_str


def clean_company_name(text):
    if not text:
        return ""
    text = re.sub(r'\([^)]*\)', '', text)
    text = re.sub(r'\bEQ\s+NEW\s+FV.*$', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\bEQ\s+FV.*$', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\bEQ\s*$', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\bF\.V\..*$', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\bMUTUAL\s+FUND\s+GOLD.*$', 'GOLD ETF',
                  text, flags=re.IGNORECASE)
    text = re.sub(r'^\d+\s+', '', text)
    text = re.sub(r'\s+', ' ', text)
    text = text.strip().upper()
    return text


# ============================================================
# STEP 3 — MAIN EXTRACTOR
# ============================================================

def extract_ais_from_pdf(pdf_file):
    text = extract_pdf_text(pdf_file)
    if not text:
        print("❌ AIS DEBUG: No text extracted from PDF")
        return None

    try:
        with open("ais_debug_extracted.txt", "w", encoding="utf-8") as f:
            f.write(text)
        print(f"✅ AIS DEBUG: Saved extracted text ({len(text)} chars)")
    except Exception as e:
        print(f"⚠️ AIS DEBUG: Could not save: {e}")

    data = parse_ais_pdf(text)
    print(f"🔍 AIS DEBUG: Extracted {len(data.get('sells', []))} sells, "
          f"{len(data.get('buys', []))} buys")

    for i, s in enumerate(data.get('sells', [])[:3]):
        print(f"   Sell {i+1}: {s['date']} | {s['company']} | "
              f"{s['isin']} | Qty={s['quantity']} | ₹{s['value']}")

    if not data['sells'] and not data['buys']:
        return None

    return data


# ============================================================
# STEP 4 — NAME COMPARISON HELPERS
# ============================================================

def normalize_company_name(name):
    if not name:
        return ""
    name = name.upper().strip()
    name = re.sub(r'[^\w\s]', ' ', name)
    words = name.split()
    cleaned = [w for w in words if w not in NOISE_WORDS]
    return ' '.join(cleaned).strip()


def similarity_score(name1, name2):
    n1 = normalize_company_name(name1)
    n2 = normalize_company_name(name2)
    if not n1 or not n2:
        return 0
    if n1 == n2:
        return 100
    return int(SequenceMatcher(None, n1, n2).ratio() * 100)


ACRONYM_SKIP_WORDS = {
    'LTD', 'LIMITED', 'PVT', 'PRIVATE', 'CO', 'COMPANY',
    'CORP', 'INC', 'THE', 'AND', '&', 'OF', 'IN', 'FOR',
    'EQ', 'NEW', 'FV', 'RS',
    '10/-', '2/-', '5/-', '1/-', 'F.V.RE.1'
}


def get_acronym(name):
    if not name:
        return ""
    words = re.sub(r'[^\w\s]', ' ', name.upper()).split()
    letters = []
    for w in words:
        if w in ACRONYM_SKIP_WORDS or not w:
            continue
        letters.append(w[0])
    return ''.join(letters)


def acronym_score(short_name, full_name):
    short = re.sub(r'[^\w]', '', short_name.upper().strip())
    if len(short) < 2:
        return 0
    full_acronym = get_acronym(full_name)
    if not full_acronym or len(full_acronym) < 2:
        return 0
    if short == full_acronym:
        return 95
    if len(short) >= 3 and short in full_acronym:
        return 85
    if len(full_acronym) >= 3 and full_acronym in short:
        return 80
    return 0


def substring_score(name_a, name_b):
    norm_a = normalize_company_name(name_a)
    norm_b = normalize_company_name(name_b)
    if not norm_a or not norm_b:
        return 0

    words_a = [w for w in norm_a.split() if len(w) >= 3]
    words_b = [w for w in norm_b.split() if len(w) >= 3]
    if not words_a or not words_b:
        return 0

    shorter, longer = (words_a, words_b) if len(words_a) <= len(words_b) \
                                          else (words_b, words_a)
    if not shorter:
        return 0

    matched = 0
    for sw in shorter:
        for lw in longer:
            if sw == lw or lw.startswith(sw) or sw.startswith(lw):
                matched += 1
                break

    if matched == 0:
        return 0

    coverage = matched / len(shorter)
    if coverage >= 1.0:
        return 85
    elif coverage >= 0.66:
        return 75
    elif coverage >= 0.5:
        return 70
    return 0


def find_candidate_matches(ais_company, app_companies_list, used_keys,
                            threshold=70):
    matches = []
    seen_keys = set()

    for app_item in app_companies_list:
        app_company = app_item.get('company', '')
        app_isin = app_item.get('isin', '')
        key = app_isin if app_isin else app_company.upper()
        if key in used_keys or key in seen_keys:
            continue

        sim_score = similarity_score(ais_company, app_company)
        acro_score = max(
            acronym_score(app_company, ais_company),
            acronym_score(ais_company, app_company)
        )
        sub_score = substring_score(ais_company, app_company)

        max_score = max(sim_score, acro_score, sub_score)

        if max_score >= threshold:
            matches.append((app_item, max_score))
            seen_keys.add(key)

    matches.sort(key=lambda x: -x[1])
    return matches


def find_fuzzy_matches(ais_company, app_companies_list, used_keys,
                       threshold=70):
    return find_candidate_matches(
        ais_company, app_companies_list, used_keys, threshold
    )


# ============================================================
# STEP 5 — AGGREGATE TRANSACTIONS
# ============================================================

def aggregate_by_company(transactions, tx_type):
    aggregated = {}
    for tx in transactions:
        isin = tx.get('isin', '')
        company = tx.get('company', '')
        key = isin if isin else company.upper()
        if key not in aggregated:
            aggregated[key] = {
                'company': company,
                'isin': isin,
                'total_qty': 0,
                'total_value': 0,
                'details': []
            }
        aggregated[key]['total_qty'] += tx.get('quantity', 0)
        aggregated[key]['total_value'] += tx.get('value', 0)
        aggregated[key]['details'].append({
            'date': tx.get('date', ''),
            'quantity': tx.get('quantity', 0),
            'value': tx.get('value', 0)
        })
    return aggregated


def aggregate_app_transactions(app_transactions):
    buys = {}
    sells = {}
    for tx in app_transactions:
        tx_type = tx['type']
        if tx_type not in ('BUY', 'SELL', 'BUYBACK'):
            continue
        company = (tx['company'] or '').upper().strip()
        try:
            isin = (tx['isin'] or '').strip().upper()
        except (KeyError, IndexError):
            isin = ''
        key = isin if isin else company
        quantity = tx['quantity'] or 0
        amount = tx['amount'] or 0
        date = tx['date'] or ''
        target = buys if tx_type == 'BUY' else sells
        if key not in target:
            target[key] = {
                'company': company,
                'isin': isin,
                'total_qty': 0,
                'total_value': 0,
                'details': []
            }
        target[key]['total_qty'] += quantity
        target[key]['total_value'] += amount
        target[key]['details'].append({
            'date': date,
            'quantity': quantity,
            'value': amount
        })
    return buys, sells


# ============================================================
# STEP 6 — RECONCILE (Phase 4 + Fast Cache)
# ============================================================

def reconcile_ais(ais_data, app_transactions, user_choices=None):
    if user_choices is None:
        user_choices = {}

    # ─── DAY 16: Build master DB cache ONCE ───
    print("📚 Building master DB cache for AIS recon...")
    master_cache = build_master_db_cache()
    print(f"📚 Loaded {len(master_cache['name_index'])} securities into cache")

    ais_buys = aggregate_by_company(ais_data.get('buys', []), 'BUY')
    ais_sells = aggregate_by_company(ais_data.get('sells', []), 'SELL')
    app_buys, app_sells = aggregate_app_transactions(app_transactions)

    app_companies_list = []
    seen = set()
    for d in (app_buys, app_sells):
        for key, info in d.items():
            if key in seen:
                continue
            seen.add(key)
            app_companies_list.append({
                'company': info['company'],
                'isin': info['isin'],
                'key': key
            })

    ais_companies_list = []
    seen = set()
    for d in (ais_buys, ais_sells):
        for key, info in d.items():
            if key in seen:
                continue
            seen.add(key)
            ais_companies_list.append({
                'company': info['company'],
                'isin': info['isin'],
                'key': key
            })

    # ─── DAY 16: Pre-compute master ISIN for each app company ───
    app_master_isins = {}
    for app_item in app_companies_list:
        master_isin = resolve_isin_from_cache(
            app_item['company'], app_item['isin'], master_cache
        )
        if master_isin:
            app_master_isins[app_item['key']] = master_isin

    matched_rows = []
    mismatched_rows = []
    ais_only_rows = []
    app_only_rows = []
    ambiguous = []

    used_ais_keys = set()
    used_app_keys = set()

    for ais_item in ais_companies_list:
        ais_key = ais_item['key']
        ais_isin = ais_item['isin']
        ais_company = ais_item['company']

        if ais_key in used_ais_keys:
            continue

        app_key = None
        match_tier = None

        # Tier 1: ISIN match
        if ais_isin:
            for app_item in app_companies_list:
                if (app_item['isin'] == ais_isin
                        and app_item['key'] not in used_app_keys):
                    app_key = app_item['key']
                    match_tier = 'isin'
                    break

        # Tier 2: Exact normalized name match
        if not app_key:
            ais_normed = normalize_company_name(ais_company)
            if ais_normed:
                for app_item in app_companies_list:
                    app_normed = normalize_company_name(app_item['company'])
                    if (app_normed == ais_normed
                            and app_item['key'] not in used_app_keys):
                        app_key = app_item['key']
                        match_tier = 'name_exact'
                        break

        # Tier 2.5: Master DB match (FAST — uses pre-computed cache)
        if not app_key:
            ais_master_isin = resolve_isin_from_cache(
                ais_company, ais_isin, master_cache
            )
            if ais_master_isin:
                for app_item in app_companies_list:
                    if app_item['key'] in used_app_keys:
                        continue
                    if app_master_isins.get(app_item['key']) == ais_master_isin:
                        app_key = app_item['key']
                        match_tier = 'master_db'
                        break

        # Tier 3: Candidate matching (acronym + substring + fuzzy)
        if not app_key:
            user_choice = user_choices.get(ais_key)
            if user_choice == 'none':
                pass
            elif user_choice:
                for app_item in app_companies_list:
                    if (app_item['key'] == user_choice
                            and app_item['key'] not in used_app_keys):
                        app_key = user_choice
                        match_tier = 'user_picked'
                        break
            else:
                candidates = find_candidate_matches(
                    ais_company, app_companies_list, used_app_keys
                )
                if len(candidates) == 1:
                    app_key = candidates[0][0]['key']
                    match_tier = f"fuzzy_{candidates[0][1]}"
                elif len(candidates) > 1:
                    ambiguous.append({
                        'ais_company': ais_company,
                        'ais_isin': ais_isin,
                        'ais_key': ais_key,
                        'candidates': [
                            {
                                'company': m[0]['company'],
                                'isin': m[0]['isin'],
                                'key': m[0]['key'],
                                'score': m[1]
                            }
                            for m in candidates[:8]
                        ]
                    })
                    continue

        row = build_comparison_row(
            ais_key, app_key,
            ais_buys, ais_sells,
            app_buys, app_sells,
            ais_company, ais_isin,
            match_tier
        )

        used_ais_keys.add(ais_key)
        if app_key:
            used_app_keys.add(app_key)

        if not app_key:
            ais_only_rows.append(row)
        elif row['is_perfect_match']:
            matched_rows.append(row)
        else:
            mismatched_rows.append(row)

    for app_item in app_companies_list:
        if app_item['key'] in used_app_keys:
            continue
        app_key = app_item['key']
        app_company = app_item['company']
        app_isin = app_item['isin']
        row = build_comparison_row(
            None, app_key,
            ais_buys, ais_sells,
            app_buys, app_sells,
            app_company, app_isin,
            None
        )
        app_only_rows.append(row)

    return {
        'matched': matched_rows,
        'mismatched': mismatched_rows,
        'ais_only': ais_only_rows,
        'app_only': app_only_rows,
        'ambiguous': ambiguous,
        'total_matched': len(matched_rows),
        'total_mismatched': len(mismatched_rows),
        'total_ais_only': len(ais_only_rows),
        'total_app_only': len(app_only_rows),
        'total_ambiguous': len(ambiguous),
        'has_ambiguous': len(ambiguous) > 0
    }


def build_comparison_row(ais_key, app_key, ais_buys, ais_sells,
                          app_buys, app_sells, company, isin, match_tier):
    ais_buy = ais_buys.get(ais_key, {}) if ais_key else {}
    ais_sell = ais_sells.get(ais_key, {}) if ais_key else {}
    app_buy = app_buys.get(app_key, {}) if app_key else {}
    app_sell = app_sells.get(app_key, {}) if app_key else {}

    app_buy_qty = app_buy.get('total_qty', 0)
    ais_buy_qty = ais_buy.get('total_qty', 0)
    app_buy_val = app_buy.get('total_value', 0)
    ais_buy_val = ais_buy.get('total_value', 0)

    app_sell_qty = app_sell.get('total_qty', 0)
    ais_sell_qty = ais_sell.get('total_qty', 0)
    app_sell_val = app_sell.get('total_value', 0)
    ais_sell_val = ais_sell.get('total_value', 0)

    buy_qty_diff = app_buy_qty - ais_buy_qty
    buy_val_diff = app_buy_val - ais_buy_val
    sell_qty_diff = app_sell_qty - ais_sell_qty
    sell_val_diff = app_sell_val - ais_sell_val

    is_perfect_match = (
        buy_qty_diff == 0 and buy_val_diff == 0 and
        sell_qty_diff == 0 and sell_val_diff == 0
    )

    detail = build_date_wise_detail(
        ais_buy.get('details', []),
        app_buy.get('details', []),
        ais_sell.get('details', []),
        app_sell.get('details', [])
    )

    use_company = company if company else (
        ais_buy.get('company') or ais_sell.get('company') or
        app_buy.get('company') or app_sell.get('company') or ''
    )
    use_isin = isin if isin else (
        ais_buy.get('isin') or ais_sell.get('isin') or
        app_buy.get('isin') or app_sell.get('isin') or ''
    )

    app_company_name = (
        app_buy.get('company') or
        app_sell.get('company') or ''
    )
    ais_company_name = (
        ais_buy.get('company') or
        ais_sell.get('company') or ''
    )

    return {
        'company': use_company,
        'isin': use_isin,
        'app_company': app_company_name,
        'ais_company': ais_company_name,
        'match_tier': match_tier,
        'app_buy_qty': app_buy_qty,
        'ais_buy_qty': ais_buy_qty,
        'buy_qty_diff': buy_qty_diff,
        'app_buy_val': app_buy_val,
        'ais_buy_val': ais_buy_val,
        'buy_val_diff': buy_val_diff,
        'app_sell_qty': app_sell_qty,
        'ais_sell_qty': ais_sell_qty,
        'sell_qty_diff': sell_qty_diff,
        'app_sell_val': app_sell_val,
        'ais_sell_val': ais_sell_val,
        'sell_val_diff': sell_val_diff,
        'is_perfect_match': is_perfect_match,
        'detail': detail
    }


def build_date_wise_detail(ais_buy_dets, app_buy_dets,
                             ais_sell_dets, app_sell_dets):
    def merge_by_date(app_list, ais_list):
        by_date = {}
        for item in app_list:
            d = item['date']
            if d not in by_date:
                by_date[d] = {'app_qty': 0, 'app_val': 0,
                              'ais_qty': 0, 'ais_val': 0}
            by_date[d]['app_qty'] += item['quantity']
            by_date[d]['app_val'] += item['value']
        for item in ais_list:
            d = item['date']
            if d not in by_date:
                by_date[d] = {'app_qty': 0, 'app_val': 0,
                              'ais_qty': 0, 'ais_val': 0}
            by_date[d]['ais_qty'] += item['quantity']
            by_date[d]['ais_val'] += item['value']
        rows = []
        for d in sorted(by_date.keys()):
            r = by_date[d]
            rows.append({
                'date': d,
                'app_qty': r['app_qty'],
                'ais_qty': r['ais_qty'],
                'qty_diff': r['app_qty'] - r['ais_qty'],
                'app_val': r['app_val'],
                'ais_val': r['ais_val'],
                'val_diff': r['app_val'] - r['ais_val']
            })
        return rows

    return {
        'buys': merge_by_date(app_buy_dets, ais_buy_dets),
        'sells': merge_by_date(app_sell_dets, ais_sell_dets)
    }


# ============================================================
# STEP 7 — EXPORT TO EXCEL
# ============================================================

def export_ais_reconciliation_to_excel(result, client_name, fin_year):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from datetime import datetime

    wb = Workbook()

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50",
                               fill_type="solid")
    bold_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center',
                       wrap_text=True)

    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA",
                             fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA",
                           fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD",
                              fill_type="solid")
    orange_fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4",
                              fill_type="solid")

    ws = wb.active
    ws.title = "Summary"
    ws['A1'] = "AIS RECONCILIATION REPORT"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = center

    ws['A3'] = "Client:"
    ws['B3'] = client_name
    ws['A4'] = "Financial Year:"
    ws['B4'] = fin_year
    ws['A5'] = "Generated:"
    ws['B5'] = datetime.now().strftime("%d-%b-%Y %H:%M")
    for c in ['A3', 'A4', 'A5']:
        ws[c].font = bold_font

    ws['A7'] = "SUMMARY"
    ws['A7'].font = Font(bold=True, size=13)

    pairs = [
        ('A9', 'B9', "✅ Perfect Match:", result['total_matched'], green_fill),
        ('A10', 'B10', "⚠️ Mismatched:", result['total_mismatched'], yellow_fill),
        ('A11', 'B11', "❌ In AIS Only:", result['total_ais_only'], orange_fill),
        ('A12', 'B12', "❌ In App Only:", result['total_app_only'], red_fill),
    ]
    for a, b, lbl, val, fill in pairs:
        ws[a] = lbl
        ws[b] = val
        ws[a].fill = fill
        ws[b].fill = fill
        ws[a].font = bold_font
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 20

    sheet_configs = [
        ('Matched', result['matched'], green_fill),
        ('Mismatched', result['mismatched'], yellow_fill),
        ('AIS Only', result['ais_only'], orange_fill),
        ('App Only', result['app_only'], red_fill),
    ]

    headers = [
        "Sr", "Company", "ISIN",
        "App Buy Qty", "AIS Buy Qty", "Buy Qty Diff",
        "App Buy Val", "AIS Buy Val", "Buy Val Diff",
        "App Sell Qty", "AIS Sell Qty", "Sell Qty Diff",
        "App Sell Val", "AIS Sell Val", "Sell Val Diff"
    ]

    for sheet_name, rows, fill in sheet_configs:
        ws = wb.create_sheet(sheet_name)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = thin_border

        for idx, row in enumerate(rows, 1):
            r = idx + 1
            ws.cell(row=r, column=1, value=idx).border = thin_border
            ws.cell(row=r, column=2, value=row['company']).border = thin_border
            ws.cell(row=r, column=3, value=row['isin']).border = thin_border
            ws.cell(row=r, column=4, value=row['app_buy_qty']).border = thin_border
            ws.cell(row=r, column=5, value=row['ais_buy_qty']).border = thin_border
            c6 = ws.cell(row=r, column=6, value=row['buy_qty_diff'])
            c6.border = thin_border
            if row['buy_qty_diff'] != 0:
                c6.fill = yellow_fill
            ws.cell(row=r, column=7, value=row['app_buy_val']).border = thin_border
            ws.cell(row=r, column=8, value=row['ais_buy_val']).border = thin_border
            c9 = ws.cell(row=r, column=9, value=row['buy_val_diff'])
            c9.border = thin_border
            if row['buy_val_diff'] != 0:
                c9.fill = yellow_fill
            ws.cell(row=r, column=10, value=row['app_sell_qty']).border = thin_border
            ws.cell(row=r, column=11, value=row['ais_sell_qty']).border = thin_border
            c12 = ws.cell(row=r, column=12, value=row['sell_qty_diff'])
            c12.border = thin_border
            if row['sell_qty_diff'] != 0:
                c12.fill = yellow_fill
            ws.cell(row=r, column=13, value=row['app_sell_val']).border = thin_border
            ws.cell(row=r, column=14, value=row['ais_sell_val']).border = thin_border
            c15 = ws.cell(row=r, column=15, value=row['sell_val_diff'])
            c15.border = thin_border
            if row['sell_val_diff'] != 0:
                c15.fill = yellow_fill

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 36
        ws.column_dimensions['C'].width = 16
        for col_letter in ['D', 'E', 'F', 'G', 'H', 'I',
                            'J', 'K', 'L', 'M', 'N', 'O']:
            ws.column_dimensions[col_letter].width = 14

    return wb