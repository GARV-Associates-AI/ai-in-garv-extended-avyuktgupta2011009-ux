# itr_generator.py
# ITR Schedule CG Generator — Day 14
# All shares treated as Listed Equity (STT Paid)
# Two Excel options: Detailed vs Adjusted
# Exemption shown ONLY in TOTAL row, not per row

from database import get_client, get_transactions, get_fmv, get_available_losses
from calculator import run_fifo, calculate_tax_summary
from datetime import datetime


# ─────────────────────────────────────────────────────────
# TAX RULES PER FINANCIAL YEAR
# ─────────────────────────────────────────────────────────

def get_tax_rules(fin_year):
    if fin_year == "2024-25":
        return {
            "stcg_111A_rate"     : 20.0,
            "ltcg_112A_rate"     : 12.5,
            "ltcg_112A_exemption": 125000,
            "ltcg_112_rate"      : 20.0,
        }
    return {
        "stcg_111A_rate"     : 15.0,
        "ltcg_112A_rate"     : 10.0,
        "ltcg_112A_exemption": 100000,
        "ltcg_112_rate"      : 20.0,
    }


def get_ay(fin_year):
    try:
        start = int(fin_year.split("-")[0])
        return f"AY {start + 1}-{str(start + 2)[-2:]}"
    except Exception:
        return ""


# ─────────────────────────────────────────────────────────
# MAIN FUNCTION
# ─────────────────────────────────────────────────────────

def generate_schedule_cg(client_id, fin_year):

    transactions = get_transactions(client_id, fin_year)
    if not transactions:
        return {"error": f"No transactions found for FY {fin_year}."}

    sells = [t for t in transactions if t["type"] in ["SELL", "BUYBACK"]]
    if not sells:
        return {"error": "No SELL or BUYBACK transactions found. Nothing to report."}

    fmv_data         = get_fmv(client_id)
    output_rows, errors = run_fifo(transactions, fmv_data)

    if not output_rows:
        return {"error": "Calculation produced no results. Check your transactions."}

    available_losses = get_available_losses(client_id, fin_year)

    tax_summary = calculate_tax_summary(
        output_rows,
        brought_forward_losses=available_losses if available_losses else None,
        apply_bf=bool(available_losses)
    )

    rules   = get_tax_rules(fin_year)
    buckets = _build_buckets(output_rows, rules)
    _compute_tax(buckets, rules)

    bf_setoff = {
        "stcl_used"     : tax_summary.get("bf_stcg_applied", 0),
        "ltcl_used"     : tax_summary.get("bf_ltcg_applied", 0),
        "stcl_remaining": tax_summary.get("stcl_carryforward", 0),
        "ltcl_remaining": tax_summary.get("ltcl_carryforward", 0),
        "log"           : tax_summary.get("bf_application_log", []),
    }

    summary = _build_summary(buckets, bf_setoff)
    client  = get_client(client_id)

    return {
        "client_name"   : client[1] if client else "Unknown",
        "client_pan"    : client[2] if client else "N/A",
        "fin_year"      : fin_year,
        "ay_year"       : get_ay(fin_year),
        "generated_on"  : datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "buckets"       : buckets,
        "bf_loss_setoff": bf_setoff,
        "summary"       : summary,
        "calc_errors"   : errors,
        "output_rows"   : output_rows,
    }


# ─────────────────────────────────────────────────────────
# BUILD BUCKETS
# ─────────────────────────────────────────────────────────

def _build_buckets(output_rows, rules):

    buckets = {
        "stcg_111A": {
            "label"       : "STCG — Listed Equity, STT Paid (Sec 111A)",
            "rate_display": f"{rules['stcg_111A_rate']}%",
            "transactions": [],
            "total_sale"  : 0.0,
            "total_cost"  : 0.0,
            "total_gain"  : 0.0,
        },
        "ltcg_112A": {
            "label"       : "LTCG — Listed Equity, STT Paid (Sec 112A)",
            "rate_display": f"{rules['ltcg_112A_rate']}%",
            "transactions": [],
            "total_sale"  : 0.0,
            "total_cost"  : 0.0,
            "total_gain"  : 0.0,
        },
    }

    for row in output_rows:
        gain_type = row.get("gain_type", "STCG")
        gain      = row.get("profit_loss", 0) or 0
        sale_val  = row.get("adjusted_proceeds", 0) or 0
        cost_val  = row.get("final_cost_used", 0) or 0

        clean = {
            "company"       : row.get("company", ""),
            "isin"          : row.get("isin", ""),
            "buy_date"      : row.get("buy_date", ""),
            "sell_date"     : row.get("sell_date", ""),
            "qty"           : row.get("shares", 0),
            "sale_value"    : round(sale_val, 2),
            "cost"          : round(cost_val, 2),
            "gain"          : round(gain, 2),
            "term"          : "Short Term" if gain_type == "STCG" else "Long Term",
            "stt"           : round(row.get("stt_paid", 0) or 0, 2),
            "holding_days"  : row.get("holding_days", 0),
            "grandfathering": row.get("grandfathering", False),
            "fmv"           : row.get("fmv_31_jan_2018", None),
            "notes"         : row.get("notes", ""),
        }

        key = "stcg_111A" if gain_type == "STCG" else "ltcg_112A"

        buckets[key]["transactions"].append(clean)
        buckets[key]["total_sale"] += sale_val
        buckets[key]["total_cost"] += cost_val
        buckets[key]["total_gain"] += gain

    for key in buckets:
        buckets[key]["total_sale"] = round(buckets[key]["total_sale"], 2)
        buckets[key]["total_cost"] = round(buckets[key]["total_cost"], 2)
        buckets[key]["total_gain"] = round(buckets[key]["total_gain"], 2)

    return buckets


# ─────────────────────────────────────────────────────────
# COMPUTE TAX
# ─────────────────────────────────────────────────────────

def _compute_tax(buckets, rules):

    exemption = rules["ltcg_112A_exemption"]

    g = buckets["stcg_111A"]["total_gain"]
    buckets["stcg_111A"]["tax"]       = round(max(0, g) * rules["stcg_111A_rate"] / 100, 2)
    buckets["stcg_111A"]["exemption"] = 0

    g112A   = buckets["ltcg_112A"]["total_gain"]
    taxable = max(0, g112A - exemption)
    buckets["ltcg_112A"]["gain_before_ex"] = round(g112A, 2)
    buckets["ltcg_112A"]["exemption"]      = exemption
    buckets["ltcg_112A"]["taxable_gain"]   = round(taxable, 2)
    buckets["ltcg_112A"]["tax"]            = round(
        taxable * rules["ltcg_112A_rate"] / 100, 2
    )


# ─────────────────────────────────────────────────────────
# BUILD SUMMARY
# ─────────────────────────────────────────────────────────

def _build_summary(buckets, bf_setoff):

    total_stcg = round(buckets["stcg_111A"]["total_gain"], 2)
    total_ltcg = round(buckets["ltcg_112A"]["total_gain"], 2)
    total_gain = round(total_stcg + total_ltcg, 2)

    stcl      = bf_setoff.get("stcl_used", 0)
    ltcl      = bf_setoff.get("ltcl_used", 0)
    net_stcg  = round(total_stcg - stcl, 2)
    net_ltcg  = round(total_ltcg - ltcl, 2)
    net_total = round(net_stcg + net_ltcg, 2)

    tax_111A  = buckets["stcg_111A"].get("tax") or 0
    tax_112A  = buckets["ltcg_112A"].get("tax") or 0

    return {
        "total_stcg"            : total_stcg,
        "total_ltcg"            : total_ltcg,
        "total_gain"            : total_gain,
        "net_stcg_after_setoff" : net_stcg,
        "net_ltcg_after_setoff" : net_ltcg,
        "net_total_after_setoff": net_total,
        "total_tax_excl_slab"   : round(tax_111A + tax_112A, 2),
    }


# ─────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────

def export_schedule_cg_to_excel(schedule, mode="detailed"):

    import io
    from openpyxl import Workbook

    wb      = Workbook()
    ws_stcg = wb.active
    ws_stcg.title = "STCG - Sec 111A"
    ws_ltcg = wb.create_sheet("LTCG - Sec 112A")
    ws_sum  = wb.create_sheet("Summary")
    ws_note = wb.create_sheet("ITR Filing Notes")

    _write_itr_sheet(
        ws_stcg, schedule,
        section_key="stcg_111A",
        section_label="SHORT TERM CAPITAL GAINS — Sec 111A (Listed Equity, STT Paid)",
        mode=mode,
    )
    _write_itr_sheet(
        ws_ltcg, schedule,
        section_key="ltcg_112A",
        section_label="LONG TERM CAPITAL GAINS — Sec 112A (Listed Equity, STT Paid)",
        mode=mode,
    )
    _write_summary_sheet(ws_sum, schedule)
    _write_notes_sheet(ws_note)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────
# WRITE ONE ITR DATA SHEET (STCG or LTCG)
# ─────────────────────────────────────────────────────────

def _write_itr_sheet(ws, schedule, section_key, section_label, mode):

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def fill(c):
        return PatternFill("solid", fgColor=c)

    def border():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def right():
        return Alignment(horizontal="right", vertical="center")

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Title ──
    ws.merge_cells("A1:S1")
    ws["A1"].value     = f"SCHEDULE CG — {section_label}"
    ws["A1"].font      = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill      = fill("000000")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:S2")
    ws["A2"].value = (
        f"Client: {schedule.get('client_name','N/A')}   |   "
        f"PAN: {schedule.get('client_pan','N/A')}   |   "
        f"FY: {schedule.get('fin_year','N/A')}   |   "
        f"{schedule.get('ay_year','')}   |   "
        f"Mode: {'Detailed' if mode=='detailed' else 'Adjusted'}"
    )
    ws["A2"].font      = Font(bold=True, size=10, color="FFFFFF")
    ws["A2"].fill      = fill("404040")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 20

    # ── Column Headers ──
    headers = [
        ("A", "S.No.",                                       6),
        ("B", "Date of\nAcquisition",                       14),
        ("C", "Name of Share",                              28),
        ("D", "ISIN Code",                                  14),
        ("E", "Whether\nSTT Charged",                       12),
        ("F", "Share\nType",                                10),
        ("G", "Actual Cost of\nAcquisition (₹)",            18),
        ("H", "FMV per share\nas on 31-Jan-2018 (₹)",      18),
        ("I", "Date of\nSale",                              14),
        ("J", "No. of\nShares Sold",                        12),
        ("K", "Consideration\nper share (₹)",               16),
        ("L", "Full Value\nConsideration (₹)",              18),
        ("M", "Expenses on\nTransfer (₹)",                  14),
        ("N", "Net\nConsideration (₹)",                     18),
        ("O", "Total FMV\n55(2)(ac)\n= J x H (₹)",         18),
        ("P", "Cost of\nAcquisition\n(Grandfathered) (₹)", 18),
        ("Q", "Exemption\nAmount (₹)",                      14),
        ("R", "Exemption\nSection",                         12),
        ("S", "Net Taxable\nCapital Gain (₹)",              18),
    ]

    for col_letter, header_text, width in headers:
        cell            = ws[f"{col_letter}4"]
        cell.value      = header_text
        cell.font       = Font(bold=True, size=9, color="FFFFFF")
        cell.fill       = fill("1a237e")
        cell.alignment  = center()
        cell.border     = border()
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[4].height = 45
    ws.freeze_panes = "A5"

    # ── Filter rows ──
    output_rows = schedule.get("output_rows", [])
    if section_key == "stcg_111A":
        rows = [r for r in output_rows if r.get("gain_type") == "STCG"]
    else:
        rows = [r for r in output_rows if r.get("gain_type") == "LTCG"]

    # ── Running totals ──
    data_row        = 5
    total_G         = 0.0
    total_L_raw     = 0.0
    total_M         = 0.0
    total_N         = 0.0
    total_O         = 0.0
    total_P         = 0.0
    total_S_pretax  = 0.0   # sum of (N - P) before exemption

    # ── Data rows ──
    for sr, row in enumerate(rows, 1):
        qty           = float(row.get("shares", 0) or 0)
        sell_unit     = float(row.get("sell_price_per_share", 0) or 0)
        adjusted_cost = float(row.get("adjusted_cost", 0) or 0)
        adjusted_net  = float(row.get("adjusted_proceeds", 0) or 0)
        sell_exp      = float(row.get("sell_expenses_portion", 0) or 0)
        fmv           = row.get("fmv_31_jan_2018", None)
        final_cost    = float(row.get("final_cost_used", 0) or 0)

        # ── G — always adjusted buy amount ──
        col_G = round(adjusted_cost, 2)

        # ── H — FMV per share ──
        col_H = round(float(fmv), 2) if fmv is not None else ""

        # ── O — Total FMV ──
        col_O = round(qty * float(fmv), 2) if fmv is not None else ""

        # ── P — Grandfathered cost (from calculator) ──
        col_P = round(final_cost, 2)

        # ── K, L, M, N based on mode ──
        if mode == "detailed":
            col_K = round(sell_unit, 2)
            col_L = round(sell_unit * qty, 2)
            col_M = round(sell_exp, 2)
            col_N = round(col_L - col_M, 2)
        else:
            col_K = round(adjusted_net / qty, 2) if qty > 0 else 0
            col_L = round(adjusted_net, 2)
            col_M = ""
            col_N = round(adjusted_net, 2)

        # ── Q, R — BLANK per row ──
        col_Q = ""
        col_R = ""

        # ── S — row level = N - P (before exemption) ──
        col_S = round(col_N - col_P, 2)

        # ── Write values ──
        values = [
            sr,
            row.get("buy_date", ""),
            row.get("company", ""),
            row.get("isin", ""),
            "Yes",
            "Listed",
            col_G,
            col_H,
            row.get("sell_date", ""),
            qty,
            col_K,
            col_L,
            col_M,
            col_N,
            col_O,
            col_P,
            col_Q,
            col_R,
            col_S,
        ]

        for col_idx, val in enumerate(values, 1):
            cell        = ws.cell(row=data_row, column=col_idx, value=val)
            cell.border = border()
            cell.alignment = left()

            if col_idx in (7, 8, 11, 12, 13, 14, 15, 16, 19):
                if isinstance(val, (int, float)):
                    cell.number_format = u'₹#,##,##0.00'
                    cell.alignment     = right()

            if col_idx == 10:
                cell.number_format = "0.00"
                cell.alignment     = right()

            if col_idx == 19 and isinstance(val, (int, float)) and val < 0:
                cell.font = Font(color="CC0000", bold=True)

        # ── Accumulate totals ──
        total_G        += col_G
        total_L_raw    += col_L if isinstance(col_L, (int, float)) else 0
        total_M        += col_M if isinstance(col_M, (int, float)) else 0
        total_N        += col_N if isinstance(col_N, (int, float)) else 0
        total_O        += col_O if isinstance(col_O, (int, float)) else 0
        total_P        += col_P
        total_S_pretax += col_S

        data_row += 1

    # ── TOTAL ROW ──
    if not rows:
        return

    # Merge A to F for "TOTAL" label
    ws.merge_cells(
        start_row=data_row, start_column=1,
        end_row=data_row,   end_column=6
    )
    tc            = ws.cell(row=data_row, column=1, value="TOTAL")
    tc.font       = Font(bold=True, size=10)
    tc.fill       = fill("D8D8D8")
    tc.alignment  = center()
    tc.border     = border()

    for col_idx in range(2, 7):
        cell        = ws.cell(row=data_row, column=col_idx)
        cell.fill   = fill("D8D8D8")
        cell.border = border()

    # ── Exemption for TOTAL row (LTCG only) ──
    total_Q = ""
    total_R = ""
    total_S = round(total_S_pretax, 2)

    if section_key == "ltcg_112A":
        ltcg_bkt    = schedule.get("buckets", {}).get("ltcg_112A", {})
        gain_before = float(ltcg_bkt.get("gain_before_ex", 0) or 0)
        exemption   = float(ltcg_bkt.get("exemption", 0) or 0)
        taxable     = float(ltcg_bkt.get("taxable_gain", 0) or 0)

        if gain_before > 0:
            total_Q = round(min(exemption, gain_before), 2)
            total_R = "112A"
            total_S = round(taxable, 2)

    # ── Write total columns ──
    total_map = {
        7 : round(total_G, 2),
        12: round(total_L_raw, 2),
        13: round(total_M, 2) if mode == "detailed" else "",
        14: round(total_N, 2),
        15: round(total_O, 2) if total_O else "",
        16: round(total_P, 2),
        17: total_Q,
        18: total_R,
        19: total_S,
    }

    for col_idx in range(7, 20):
        cell        = ws.cell(row=data_row, column=col_idx)
        cell.fill   = fill("D8D8D8")
        cell.border = border()
        cell.font   = Font(bold=True, size=10)

        val = total_map.get(col_idx, "")
        if val == "":
            continue

        cell.value = val

        if col_idx in (7, 12, 13, 14, 15, 16, 17, 19):
            if isinstance(val, (int, float)):
                cell.number_format = u'₹#,##,##0.00'
                cell.alignment     = right()

        if col_idx == 18:
            cell.alignment = center()

        if col_idx == 19 and isinstance(val, (int, float)) and val < 0:
            cell.font = Font(bold=True, color="CC0000")


# ─────────────────────────────────────────────────────────
# SUMMARY SHEET
# ─────────────────────────────────────────────────────────

def _write_summary_sheet(ws, schedule):

    from openpyxl.styles import Font, PatternFill, Alignment

    def fill(c):
        return PatternFill("solid", fgColor=c)

    def center():
        return Alignment(horizontal="center", vertical="center")

    def left(indent=0):
        return Alignment(horizontal="left", vertical="center", indent=indent)

    def right():
        return Alignment(horizontal="right", vertical="center")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 2

    ws.merge_cells("A1:D1")
    ws["A1"].value     = "SCHEDULE CG — SUMMARY"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = fill("000000")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:D2")
    ws["A2"].value = (
        f"Client: {schedule.get('client_name','N/A')}   |   "
        f"PAN: {schedule.get('client_pan','N/A')}   |   "
        f"FY: {schedule.get('fin_year','N/A')}   |   "
        f"{schedule.get('ay_year','')}"
    )
    ws["A2"].font      = Font(bold=True, size=10, color="FFFFFF")
    ws["A2"].fill      = fill("404040")
    ws["A2"].alignment = center()

    row     = 4
    buckets = schedule.get("buckets", {})

    def section_header(label):
        nonlocal row
        ws.merge_cells(f"A{row}:D{row}")
        c            = ws[f"A{row}"]
        c.value      = f"  {label}"
        c.font       = Font(bold=True, color="FFFFFF", size=10)
        c.fill       = fill("1a237e")
        c.alignment  = left(1)
        row += 1

    def field(label, value, bold=False):
        nonlocal row
        ws[f"B{row}"].value     = label
        ws[f"B{row}"].font      = Font(bold=bold, size=10)
        ws[f"B{row}"].alignment = left(2)
        vc                      = ws[f"C{row}"]
        vc.value                = round(float(value), 2) if value is not None else 0
        vc.number_format        = u'₹#,##,##0.00'
        vc.alignment            = right()
        if bold:
            vc.font              = Font(bold=True, size=10)
            ws[f"B{row}"].fill  = fill("E8E8E8")
            vc.fill             = fill("E8E8E8")
        row += 1

    bkt = buckets.get("stcg_111A", {})
    section_header(bkt.get("label", "STCG — Sec 111A"))
    field("Sale Proceeds", bkt.get("total_sale", 0))
    field("Cost of Acquisition", bkt.get("total_cost", 0))
    field("STCG (Sec 111A)", bkt.get("total_gain", 0), bold=True)
    field(f"Tax @ {bkt.get('rate_display','')}", bkt.get("tax", 0))
    row += 1

    bkt = buckets.get("ltcg_112A", {})
    section_header(bkt.get("label", "LTCG — Sec 112A"))
    field("Sale Proceeds", bkt.get("total_sale", 0))
    field("Cost of Acquisition", bkt.get("total_cost", 0))
    field("LTCG before exemption", bkt.get("gain_before_ex", 0), bold=True)
    field(
        f"Less: Exemption u/s 112A (Rs.{bkt.get('exemption',0):,.0f})",
        bkt.get("exemption", 0)
    )
    field("Taxable LTCG (Sec 112A)", bkt.get("taxable_gain", 0), bold=True)
    field(f"Tax @ {bkt.get('rate_display','')}", bkt.get("tax", 0))
    row += 1

    bf = schedule.get("bf_loss_setoff", {})
    if bf.get("stcl_used", 0) or bf.get("ltcl_used", 0):
        section_header("B/F LOSS SET-OFF")
        if bf.get("stcl_used", 0):
            field("B/F STCL Set Off", bf["stcl_used"])
        if bf.get("ltcl_used", 0):
            field("B/F LTCL Set Off", bf["ltcl_used"])
        row += 1

    s = schedule.get("summary", {})
    section_header("GRAND SUMMARY")
    field("Total STCG", s.get("total_stcg", 0))
    field("Total LTCG", s.get("total_ltcg", 0))
    field("Total Capital Gains", s.get("total_gain", 0), bold=True)
    row += 1
    field("Net STCG after B/F Loss", s.get("net_stcg_after_setoff", 0))
    field("Net LTCG after B/F Loss", s.get("net_ltcg_after_setoff", 0))
    field("NET CAPITAL GAINS", s.get("net_total_after_setoff", 0), bold=True)


# ─────────────────────────────────────────────────────────
# NOTES SHEET
# ─────────────────────────────────────────────────────────

def _write_notes_sheet(ws):

    from openpyxl.styles import Font

    ws.column_dimensions["A"].width = 80

    notes = [
        ("ITR SCHEDULE CG — FILING GUIDE", True, 13),
        ("", False, 10),
        ("All shares treated as: Listed Equity, STT Paid", True, 11),
        ("", False, 10),
        ("1. STCG — Sec 111A", True, 10),
        ("   -> Schedule CG -> Part A -> Row A1", False, 10),
        ("   -> Tax: 15% (up to FY 2023-24)  |  20% (FY 2024-25 onwards)", False, 10),
        ("", False, 10),
        ("2. LTCG — Sec 112A", True, 10),
        ("   -> Schedule CG -> Part B -> Row B5", False, 10),
        ("   -> Exemption: Rs.1,00,000 (up to FY 2023-24)  |  Rs.1,25,000 (FY 2024-25)", False, 10),
        ("   -> Tax: 10% (up to FY 2023-24)  |  12.5% (FY 2024-25 onwards)", False, 10),
        ("   -> FMV on 31-Jan-2018 used for pre-2018 shares", False, 10),
        ("", False, 10),
        ("EXEMPTION (Column Q):", True, 10),
        ("   Exemption is shown ONLY in the TOTAL row, not per individual row.", False, 10),
        ("   Column S per row = N - P (before exemption).", False, 10),
        ("   TOTAL row Column S = Taxable LTCG after exemption.", False, 10),
        ("", False, 10),
        ("TWO EXCEL MODES:", True, 10),
        ("   Detailed: Sell expenses shown separately in Column M", False, 10),
        ("   Adjusted: Sell expenses absorbed into sale price, Column M blank", False, 10),
        ("", False, 10),
        ("IMPORTANT:", True, 10),
        ("   STT is NOT deductible as expense u/s 40(a)(ib)", False, 10),
        ("   Bonus shares: Cost = Rs.0 as per Sec 55(2)(aa)", False, 10),
        ("   Gift/Inherited: original owner's cost and date used", False, 10),
        ("   B/F losses expire after 8 assessment years", False, 10),
        ("   File ITR before due date to carry forward losses", False, 10),
    ]

    for r, (text, bold, size) in enumerate(notes, 1):
        cell      = ws.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, size=size)