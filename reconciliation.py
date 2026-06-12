"""
reconciliation.py — Day 16 (Phase 4: Master DB integration + Fast Cache)
DP Statement Reconciliation Tool

Matching strategy:
  Tier 1: ISIN match           → AUTO ACCEPT
  Tier 2: Exact normalized name → AUTO ACCEPT
  Tier 2.5: Master DB match     → AUTO ACCEPT (fast cache)
  Tier 3: Fuzzy 70%+:
            - 1 candidate  → AUTO ACCEPT
            - 2+ candidates → ASK USER (ambiguous)
"""

import re
import json
import time
from datetime import datetime
from difflib import SequenceMatcher
from google import genai
from config import GEMINI_API_KEY
import pypdf
from openpyxl import load_workbook

# Day 16 — Fast master DB cache
from database import build_master_db_cache, resolve_isin_from_cache


# ============================================================
# NOISE WORDS
# ============================================================

NOISE_WORDS = {
    'LTD', 'LIMITED', 'PVT', 'PRIVATE', 'CO', 'COMPANY',
    'CORP', 'CORPORATION', 'INC', 'THE', 'AND', '&',
    'EQ', 'NEW', 'FV', 'RS', 'INDIA', 'INDIAN',
    'GROUP', 'HOLDINGS', 'INDUSTRIES', 'ENTERPRISES',
    'INTERNATIONAL', 'GLOBAL', '10/-', '2/-', '5/-', '1/-',
    'OF', 'IN', 'FOR'
}


# ============================================================
# STEP 1 — EXTRACT PDF TEXT
# ============================================================

def extract_pdf_text(pdf_file):
    try:
        reader = pypdf.PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"PDF read error: {e}")
        return None


# ============================================================
# STEP 2 — SMART RULES PARSER
# ============================================================

def parse_holdings_smart(text):
    bill_date = None
    date_match = re.search(
        r'Holdings?\s+as\s+on\s+(\d{1,2}[/-]\d{1,2}[/-]\d{4})',
        text, re.IGNORECASE
    )
    if date_match:
        bill_date = normalize_date(date_match.group(1))
    else:
        date_match = re.search(
            r'(?:Statement|Bill)\s+Date[\s:]+(\d{1,2}[/-]\d{1,2}[/-]\d{4})',
            text, re.IGNORECASE
        )
        if date_match:
            bill_date = normalize_date(date_match.group(1))

    holdings_start = -1
    m = re.search(r'Holdings?\s+as\s+on\s+\d{1,2}[/-]\d{1,2}[/-]\d{4}',
                  text, re.IGNORECASE)
    if m:
        holdings_start = m.end()

    if holdings_start > 0:
        holdings_text = text[holdings_start:]
    else:
        holdings_text = text

    holdings = []
    seen_isins = set()
    lines = holdings_text.split('\n')

    line_pattern = re.compile(
        r'^(.*?)(IN[EFA9][A-Z0-9]{9})(\d{1,4})\s+(.+?)\s+([\d,]+\.\d+)\s*$'
    )

    for line in lines:
        line = line.strip()
        if not line:
            continue

        match = line_pattern.match(line)
        if not match:
            continue

        numbers_before = match.group(1).strip()
        isin = match.group(2)
        company = match.group(4).strip()

        if isin in seen_isins:
            continue

        nums = re.findall(r'[\d,]+\.\d+|\d+', numbers_before)
        clean_numbers = []
        for n in nums:
            try:
                clean_numbers.append(float(n.replace(',', '')))
            except ValueError:
                pass

        if len(clean_numbers) < 2:
            continue

        free_bal = clean_numbers[0]
        pldg_bal = 0
        if len(clean_numbers) >= 3:
            for mid_num in clean_numbers[1:-1]:
                pldg_bal += mid_num

        company = company.upper().strip()
        company = re.sub(r'\s+', ' ', company)

        if len(company) < 2:
            continue

        total_qty = free_bal + pldg_bal
        if total_qty <= 0:
            continue

        seen_isins.add(isin)
        holdings.append({
            'isin': isin,
            'company': company,
            'free_qty': free_bal,
            'pledged_qty': pldg_bal,
            'total_qty': total_qty
        })

    if not holdings:
        return bill_date, None

    return bill_date, holdings


def normalize_date(date_str):
    try:
        date_str = date_str.replace('-', '/')
        parts = date_str.split('/')
        if len(parts) == 3:
            d, m, y = parts
            return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    except Exception:
        pass
    return date_str


# ============================================================
# STEP 3 — AI FALLBACK (GEMINI)
# ============================================================

def parse_with_ai(text):
    prompt = f"""You are reading a DP (Depository) holding statement from an Indian stock broker.

IMPORTANT: The PDF may contain BOTH transaction history AND a holdings snapshot.
Only extract from the HOLDINGS section (look for headers like
"Holding as on DD/MM/YYYY"). IGNORE individual transaction entries.

Extract and return as JSON:

1. bill_date — Format: YYYY-MM-DD
2. holdings — Array of:
   - "isin" — 12 chars starting INE/IN9/INF
   - "company" — UPPERCASE
   - "free_qty" — number
   - "pledged_qty" — number (0 if none)
   - "total_qty" — sum of free + pledged

Skip 0-qty rows. Return ONLY valid JSON.

DP Statement text:
{text[:18000]}

Return ONLY the JSON object."""

    backoff_seconds = [3, 6, 10, 15, 20, 30, 45, 60, 90, 120]

    for attempt in range(len(backoff_seconds)):
        try:
            client = genai.Client(api_key=GEMINI_API_KEY)
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=prompt
            )

            result_text = response.text.strip()
            if result_text.startswith("```"):
                result_text = re.sub(r'^```(?:json)?\s*', '', result_text)
                result_text = re.sub(r'\s*```$', '', result_text)

            data = json.loads(result_text)
            bill_date = data.get('bill_date')
            holdings = data.get('holdings', [])

            cleaned_holdings = []
            for h in holdings:
                total = h.get('total_qty') or (
                    (h.get('free_qty', 0) or 0) +
                    (h.get('pledged_qty', 0) or 0)
                )
                if total > 0:
                    cleaned_holdings.append({
                        'isin': h.get('isin', '').strip().upper(),
                        'company': h.get('company', '').strip().upper(),
                        'free_qty': h.get('free_qty', 0) or 0,
                        'pledged_qty': h.get('pledged_qty', 0) or 0,
                        'total_qty': total
                    })

            return bill_date, cleaned_holdings

        except json.JSONDecodeError as e:
            print(f"AI returned invalid JSON (attempt {attempt+1}): {e}")
            if attempt == len(backoff_seconds) - 1:
                return None, None
            time.sleep(5)

        except Exception as e:
            error_str = str(e)
            wait = backoff_seconds[attempt]
            if ("503" in error_str or "UNAVAILABLE" in error_str
                    or "overload" in error_str.lower()
                    or "RESOURCE_EXHAUSTED" in error_str
                    or "429" in error_str):
                print(f"Gemini overloaded (attempt {attempt+1}/10), "
                      f"retrying in {wait}s...")
                time.sleep(wait)
            else:
                print(f"AI error (attempt {attempt+1}/10): {e}")
                if attempt == len(backoff_seconds) - 1:
                    return None, None
                time.sleep(wait)

    return None, None


# ============================================================
# STEP 4 — MAIN EXTRACTOR
# ============================================================

def extract_holdings_from_pdf(pdf_file):
    text = extract_pdf_text(pdf_file)
    if not text:
        return None, None, 'failed'

    bill_date, holdings = parse_holdings_smart(text)
    if holdings and len(holdings) > 0:
        print(f"✅ Rules parser extracted {len(holdings)} holdings")
        return bill_date, holdings, 'rules'

    print("⚠️ Rules-based parser found no holdings, trying AI...")
    bill_date, holdings = parse_with_ai(text)
    if holdings and len(holdings) > 0:
        print(f"✅ AI extracted {len(holdings)} holdings")
        return bill_date, holdings, 'ai'

    print("❌ Both rules and AI failed to extract holdings")
    return None, None, 'failed'


# ============================================================
# STEP 4B — MANUAL EXCEL UPLOAD (FALLBACK)
# ============================================================

def extract_holdings_from_excel(excel_file):
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active

        bill_date = None
        try:
            b1 = ws['B1'].value
            if b1:
                if isinstance(b1, datetime):
                    bill_date = b1.strftime("%Y-%m-%d")
                else:
                    bill_date_str = str(b1).strip()
                    bill_date = normalize_date(bill_date_str)
        except Exception:
            pass

        holdings = []
        row_num = 4

        while True:
            isin = ws.cell(row=row_num, column=1).value
            company = ws.cell(row=row_num, column=2).value
            free = ws.cell(row=row_num, column=3).value
            pledged = ws.cell(row=row_num, column=4).value

            if not isin and not company:
                break

            isin = str(isin).strip().upper() if isin else ''
            company = str(company).strip().upper() if company else ''

            try:
                free = float(free) if free else 0
            except (ValueError, TypeError):
                free = 0

            try:
                pledged = float(pledged) if pledged else 0
            except (ValueError, TypeError):
                pledged = 0

            total = free + pledged

            if total > 0 and (isin or company):
                holdings.append({
                    'isin': isin,
                    'company': company,
                    'free_qty': free,
                    'pledged_qty': pledged,
                    'total_qty': total
                })

            row_num += 1
            if row_num > 5000:
                break

        if not holdings:
            return bill_date, None, 'failed'

        return bill_date, holdings, 'manual'

    except Exception as e:
        print(f"Excel read error: {e}")
        return None, None, 'failed'


def generate_manual_holdings_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "DP Holdings"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E",
                               fill_type="solid")
    label_font = Font(bold=True, size=11)
    yellow_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1",
                              fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')

    ws['A1'] = "Bill Date (YYYY-MM-DD):"
    ws['A1'].font = label_font
    ws['B1'].fill = yellow_fill
    ws['B1'].border = border
    ws['B1'] = "2026-05-31"
    ws['C1'] = "← Fill in the date as on which holdings are shown"

    headers = ["ISIN", "Company Name", "Free Qty", "Pledged Qty"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    examples = [
        ("INE002A01018", "RELIANCE INDUSTRIES", 100, 0),
        ("INE467B01029", "TATA CONSULTANCY SERVICES", 50, 0),
        ("INE040A01034", "HDFC BANK", 200, 50),
    ]
    for idx, (isin, name, free, pledged) in enumerate(examples, 4):
        ws.cell(row=idx, column=1, value=isin).border = border
        ws.cell(row=idx, column=2, value=name).border = border
        ws.cell(row=idx, column=3, value=free).border = border
        ws.cell(row=idx, column=4, value=pledged).border = border

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("How to Use This Template", True, 14),
        ("", False, 11),
        ("1. Fill the Bill Date in cell B1 (format: YYYY-MM-DD)", False, 11),
        ("2. Delete the 3 example rows", False, 11),
        ("3. For each holding, fill: ISIN, Company, Free Qty, Pledged Qty", False, 11),
        ("4. Leave Pledged Qty as 0 if none", False, 11),
        ("5. Save and upload back to the app", False, 11),
        ("", False, 11),
        ("TIP: You can copy-paste from any DP statement", True, 11),
        ("", False, 11),
        ("ISIN format: 12 chars starting with INE/IN9/INF", False, 11),
        ("Example: INE002A01018 (Reliance)", False, 11),
    ]
    for idx, (text, bold, size) in enumerate(instructions, 1):
        c = ws2.cell(row=idx, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=size)
        else:
            c.font = Font(size=size)
    ws2.column_dimensions['A'].width = 70

    return wb


# ============================================================
# STEP 5 — COMBINE MULTIPLE DP STATEMENTS
# ============================================================

def combine_holdings(holdings_list_of_lists, sources_list):
    combined = {}

    for idx, holdings in enumerate(holdings_list_of_lists):
        source = sources_list[idx] if idx < len(sources_list) \
                 else f"PDF {idx+1}"

        for h in holdings:
            isin = h.get('isin', '')
            company = h.get('company', '').upper()
            key = isin if isin else company

            if key in combined:
                combined[key]['free_qty'] += h.get('free_qty', 0)
                combined[key]['pledged_qty'] += h.get('pledged_qty', 0)
                combined[key]['total_qty'] += h.get('total_qty', 0)
                combined[key]['sources'].append(
                    f"{source}: {h.get('total_qty', 0):.0f}"
                )
            else:
                combined[key] = {
                    'isin': isin,
                    'company': h['company'],
                    'free_qty': h.get('free_qty', 0),
                    'pledged_qty': h.get('pledged_qty', 0),
                    'total_qty': h.get('total_qty', 0),
                    'sources': [f"{source}: {h.get('total_qty', 0):.0f}"]
                }

    return list(combined.values())


# ============================================================
# STEP 6 — NAME COMPARISON HELPERS
# ============================================================

def normalize_company_name(name):
    if not name:
        return ""
    name = name.upper().strip()
    name = re.sub(r'[^\w\s]', ' ', name)
    words = name.split()
    cleaned = [w for w in words if w not in NOISE_WORDS]
    return ' '.join(cleaned).strip()


def similarity_score(name1, name2):
    n1 = normalize_company_name(name1)
    n2 = normalize_company_name(name2)
    if not n1 or not n2:
        return 0
    if n1 == n2:
        return 100
    return int(SequenceMatcher(None, n1, n2).ratio() * 100)


def find_fuzzy_matches(dp_company, app_holdings, used_app_keys,
                       threshold=70):
    matches = []
    for app_item in app_holdings:
        app_company = app_item.get('company', '')
        app_isin = app_item.get('isin', '')
        key = app_isin if app_isin else app_company.upper()
        if key in used_app_keys:
            continue
        score = similarity_score(dp_company, app_company)
        if score >= threshold:
            matches.append((app_item, score))
    matches.sort(key=lambda x: -x[1])
    return matches


# ============================================================
# STEP 7 — RECONCILE (Phase 4 + Fast Cache)
# ============================================================

def reconcile(dp_holdings, app_holdings, user_choices=None):
    if user_choices is None:
        user_choices = {}

    # ─── DAY 16: Build master DB cache ONCE for this run ───
    print("📚 Building master DB cache...")
    master_cache = build_master_db_cache()
    print(f"📚 Loaded {len(master_cache['name_index'])} securities into cache")

    matched, mismatched, dp_only, app_only, ambiguous = [], [], [], [], []
    used_app_keys = set()

    app_aggregated = {}
    for h in app_holdings:
        isin = h.get('isin', '')
        company = h.get('company', '')
        key = isin if isin else company.upper()
        if key not in app_aggregated:
            app_aggregated[key] = {
                'isin': isin,
                'company': company,
                'total_qty': 0
            }
        app_aggregated[key]['total_qty'] += h.get('quantity', 0)
    app_list = list(app_aggregated.values())

    app_by_isin = {}
    app_by_norm = {}
    for h in app_list:
        isin = h.get('isin', '')
        company = h.get('company', '')
        if isin:
            app_by_isin[isin] = h
        if company:
            normed = normalize_company_name(company)
            if normed:
                app_by_norm[normed] = h

    # ─── DAY 16: Pre-compute master ISIN for each app holding ───
    # (so we don't re-compute inside the DP loop)
    app_master_isins = {}
    for h in app_list:
        h_isin = h.get('isin', '')
        h_company = h.get('company', '')
        h_key = h_isin if h_isin else h_company.upper()
        master_isin = resolve_isin_from_cache(h_company, h_isin, master_cache)
        if master_isin:
            app_master_isins[h_key] = master_isin

    for dp_item in dp_holdings:
        dp_isin = dp_item.get('isin', '')
        dp_company = dp_item['company']
        dp_qty = dp_item['total_qty']
        dp_key = dp_isin if dp_isin else dp_company.upper()

        app_item = None
        match_tier = None

        # Tier 1: ISIN exact match
        if dp_isin and dp_isin in app_by_isin:
            cand = app_by_isin[dp_isin]
            ck = cand.get('isin', '') or cand.get('company', '').upper()
            if ck not in used_app_keys:
                app_item = cand
                match_tier = 'isin'

        # Tier 2: Normalized name exact match
        if not app_item:
            dp_n = normalize_company_name(dp_company)
            if dp_n and dp_n in app_by_norm:
                cand = app_by_norm[dp_n]
                ck = cand.get('isin', '') or cand.get('company', '').upper()
                if ck not in used_app_keys:
                    app_item = cand
                    match_tier = 'name_exact'

        # Tier 2.5: Master DB match (FAST — uses pre-computed cache)
        if not app_item:
            dp_master_isin = resolve_isin_from_cache(
                dp_company, dp_isin, master_cache
            )
            if dp_master_isin:
                for h in app_list:
                    h_isin = h.get('isin', '')
                    h_key = h_isin if h_isin else h.get('company', '').upper()
                    if h_key in used_app_keys:
                        continue
                    if app_master_isins.get(h_key) == dp_master_isin:
                        app_item = h
                        match_tier = 'master_db'
                        break

        # Tier 3: Fuzzy / user pick
        if not app_item:
            user_choice = user_choices.get(dp_key)
            if user_choice == 'none':
                pass
            elif user_choice:
                for h in app_list:
                    h_isin = h.get('isin', '')
                    h_key = h_isin if h_isin else h.get('company', '').upper()
                    if h_key == user_choice and h_key not in used_app_keys:
                        app_item = h
                        match_tier = 'user_picked'
                        break
            else:
                fuzzy = find_fuzzy_matches(dp_company, app_list, used_app_keys)
                if len(fuzzy) == 0:
                    pass
                elif len(fuzzy) == 1:
                    app_item = fuzzy[0][0]
                    match_tier = f"fuzzy_{fuzzy[0][1]}"
                else:
                    ambiguous.append({
                        'dp_company': dp_company,
                        'dp_isin': dp_isin,
                        'dp_qty': dp_qty,
                        'dp_key': dp_key,
                        'candidates': [
                            {
                                'company': m[0].get('company', ''),
                                'isin': m[0].get('isin', ''),
                                'app_qty': m[0].get('total_qty', 0),
                                'score': m[1],
                                'key': (m[0].get('isin') or
                                        m[0].get('company', '').upper())
                            }
                            for m in fuzzy[:5]
                        ]
                    })
                    continue

        if app_item:
            a_isin = app_item.get('isin', '')
            a_co = app_item.get('company', '')
            a_key = a_isin if a_isin else a_co.upper()
            used_app_keys.add(a_key)
            a_qty = app_item.get('total_qty', 0)

            row = {
                'company': dp_company,
                'isin': dp_isin or a_isin,
                'app_company': a_co,
                'dp_qty': dp_qty,
                'app_qty': a_qty,
                'difference': dp_qty - a_qty,
                'free_qty': dp_item.get('free_qty', 0),
                'pledged_qty': dp_item.get('pledged_qty', 0),
                'match_tier': match_tier
            }
            if abs(dp_qty - a_qty) < 0.01:
                matched.append(row)
            else:
                row['suggestion'] = suggest_cause(dp_qty, a_qty)
                row['cause'] = row['suggestion']
                mismatched.append(row)
        else:
            if (not user_choices.get(dp_key)
                    or user_choices.get(dp_key) == 'none'):
                dp_only.append({
                    'company': dp_company,
                    'isin': dp_isin,
                    'dp_qty': dp_qty,
                    'free_qty': dp_item.get('free_qty', 0),
                    'pledged_qty': dp_item.get('pledged_qty', 0)
                })

    for app_item in app_list:
        a_isin = app_item.get('isin', '')
        a_co = app_item.get('company', '')
        a_key = a_isin if a_isin else a_co.upper()
        if a_key in used_app_keys:
            continue
        app_only.append({
            'company': a_co,
            'isin': a_isin,
            'app_qty': app_item.get('total_qty', 0)
        })

    return {
        'matched': matched,
        'mismatched': mismatched,
        'dp_only': dp_only,
        'app_only': app_only,
        'ambiguous': ambiguous,
        'total_dp': len(dp_holdings),
        'total_app': len(app_list),
        'total_matched': len(matched),
        'total_mismatched': len(mismatched),
        'total_dp_only': len(dp_only),
        'total_app_only': len(app_only),
        'total_ambiguous': len(ambiguous),
        'has_ambiguous': len(ambiguous) > 0
    }


def suggest_cause(dp_qty, app_qty):
    diff = dp_qty - app_qty
    if app_qty == 0:
        return "⚠️ App shows zero — check for missed BUY entries"
    ratio = dp_qty / app_qty if app_qty > 0 else 0
    if ratio == 2:
        return "🔍 Likely missed 1:1 BONUS issue"
    elif ratio == 3:
        return "🔍 Likely missed 2:1 BONUS issue"
    elif ratio == 1.5:
        return "🔍 Likely missed 1:2 BONUS issue"
    elif ratio == 5 or ratio == 10:
        return f"🔍 Likely missed STOCK SPLIT ({int(ratio)}×)"
    elif diff > 0:
        return (f"➕ DP has {diff:.0f} more — "
                f"check missed BONUS / off-market IN")
    else:
        return (f"➖ App has {abs(diff):.0f} more — "
                f"check missed SELL / off-market OUT")


# ============================================================
# STEP 8 — EXPORT TO EXCEL
# ============================================================

def export_reconciliation_to_excel(result, client_name, bill_date):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50",
                               fill_type="solid")
    bold_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center',
                       wrap_text=True)

    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA",
                             fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA",
                           fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD",
                              fill_type="solid")
    orange_fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4",
                              fill_type="solid")

    ws = wb.active
    ws.title = "Summary"
    ws['A1'] = "RECONCILIATION REPORT"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = center

    ws['A3'] = "Client:"
    ws['B3'] = client_name
    ws['A4'] = "Bill Date:"
    ws['B4'] = bill_date
    ws['A5'] = "Generated:"
    ws['B5'] = datetime.now().strftime("%d-%b-%Y %H:%M")
    for c in ['A3', 'A4', 'A5']:
        ws[c].font = bold_font

    ws['A7'] = "SUMMARY"
    ws['A7'].font = Font(bold=True, size=13)
    ws['A9'] = "Total in DP Statement:"
    ws['B9'] = result['total_dp']
    ws['A10'] = "Total in App:"
    ws['B10'] = result['total_app']

    pairs = [
        ('A12', 'B12', "✅ Matched:", result['total_matched'], green_fill),
        ('A13', 'B13', "⚠️ Mismatched:", result['total_mismatched'], yellow_fill),
        ('A14', 'B14', "❌ In DP, Not in App:", result['total_dp_only'], orange_fill),
        ('A15', 'B15', "❌ In App, Not in DP:", result['total_app_only'], red_fill),
    ]
    for a, b, lbl, val, fill in pairs:
        ws[a] = lbl
        ws[b] = val
        ws[a].fill = fill
        ws[b].fill = fill
    for r in range(9, 16):
        ws[f'A{r}'].font = bold_font
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 20

    ws2 = wb.create_sheet("Matched")
    headers = ["Sr", "Company", "ISIN", "DP Qty", "App Qty", "Match Type"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    for idx, row in enumerate(result['matched'], 1):
        tier = row.get('match_tier', 'isin')
        tier_label = {
            'isin': 'ISIN',
            'name_exact': 'Name Match',
            'master_db': 'Master DB',
            'user_picked': 'User Picked'
        }.get(tier, tier.upper() if tier else 'ISIN')
        if tier and tier.startswith('fuzzy_'):
            tier_label = f"Fuzzy {tier.split('_')[1]}%"

        ws2.cell(row=idx+1, column=1, value=idx).border = thin_border
        ws2.cell(row=idx+1, column=2, value=row['company']).border = thin_border
        ws2.cell(row=idx+1, column=3, value=row['isin']).border = thin_border
        ws2.cell(row=idx+1, column=4, value=row['dp_qty']).border = thin_border
        ws2.cell(row=idx+1, column=5, value=row['app_qty']).border = thin_border
        c = ws2.cell(row=idx+1, column=6, value=tier_label)
        c.border = thin_border
        c.fill = green_fill

    for col, w in [('A', 6), ('B', 40), ('C', 16),
                   ('D', 12), ('E', 12), ('F', 16)]:
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Mismatched")
    headers = ["Sr", "Company", "ISIN", "DP Qty", "App Qty",
               "Difference", "Likely Cause"]
    for col, h in enumerate(headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    for idx, row in enumerate(result['mismatched'], 1):
        ws3.cell(row=idx+1, column=1, value=idx).border = thin_border
        ws3.cell(row=idx+1, column=2, value=row['company']).border = thin_border
        ws3.cell(row=idx+1, column=3, value=row['isin']).border = thin_border
        ws3.cell(row=idx+1, column=4, value=row['dp_qty']).border = thin_border
        ws3.cell(row=idx+1, column=5, value=row['app_qty']).border = thin_border
        c = ws3.cell(row=idx+1, column=6, value=row['difference'])
        c.border = thin_border
        c.fill = yellow_fill
        ws3.cell(row=idx+1, column=7,
                 value=row.get('suggestion', '')).border = thin_border

    for col, w in [('A', 6), ('B', 40), ('C', 16), ('D', 12),
                   ('E', 12), ('F', 14), ('G', 50)]:
        ws3.column_dimensions[col].width = w

    ws4 = wb.create_sheet("In DP Only")
    headers = ["Sr", "Company", "ISIN", "DP Qty", "Free Qty", "Pledged Qty"]
    for col, h in enumerate(headers, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    for idx, row in enumerate(result['dp_only'], 1):
        ws4.cell(row=idx+1, column=1, value=idx).border = thin_border
        ws4.cell(row=idx+1, column=2, value=row['company']).border = thin_border
        ws4.cell(row=idx+1, column=3, value=row['isin']).border = thin_border
        ws4.cell(row=idx+1, column=4, value=row['dp_qty']).border = thin_border
        ws4.cell(row=idx+1, column=5,
                 value=row.get('free_qty', 0)).border = thin_border
        ws4.cell(row=idx+1, column=6,
                 value=row.get('pledged_qty', 0)).border = thin_border

    for col, w in [('A', 6), ('B', 40), ('C', 16),
                   ('D', 12), ('E', 12), ('F', 14)]:
        ws4.column_dimensions[col].width = w

    ws5 = wb.create_sheet("In App Only")
    headers = ["Sr", "Company", "ISIN", "App Qty"]
    for col, h in enumerate(headers, 1):
        c = ws5.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    for idx, row in enumerate(result['app_only'], 1):
        ws5.cell(row=idx+1, column=1, value=idx).border = thin_border
        ws5.cell(row=idx+1, column=2, value=row['company']).border = thin_border
        ws5.cell(row=idx+1, column=3, value=row['isin']).border = thin_border
        ws5.cell(row=idx+1, column=4, value=row['app_qty']).border = thin_border

    for col, w in [('A', 6), ('B', 40), ('C', 16), ('D', 12)]:
        ws5.column_dimensions[col].width = w

    return wb