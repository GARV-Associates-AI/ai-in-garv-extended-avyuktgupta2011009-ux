# excel_handler.py
# ============================================================
# Handles Excel file import and template generation
# UPDATED: Day 10 — STT column added (informational only)
# UPDATED: Day 12 — generate_consolidated_excel added
# UPDATED: Day 12.5 — consolidated excel now lot-wise with Buy Date
# UPDATED: Day 12.5 — BUY with amount=0 warns instead of rejects
# UPDATED: Day 12.5 — Warnings now separate from errors (3-tuple return)
# UPDATED: Day 12.5 — Fixed MergedCell read-only error
# UPDATED: Day 12.5 — Consolidated reports now use Rate (with buy exp adj)
# UPDATED: Day 12.5 — Added Layout 2 (Per-Client Date) consolidated report
# UPDATED: Day 12.6 — Layout 2 FIXED:
#                     - Company name + ISIN repeat on EVERY row
#                     - Same date across clients = SAME row
#                     - Different dates = different rows
# UPDATED: Day 12.6 — Consolidated reports stripped of all colors
#                     (clean B&W look — borders + bold headers only)
# ============================================================

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


def generate_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "Date", "Type", "Company", "ISIN", "Quantity",
        "Amount", "Buy_Expenses", "Sell_Expenses", "STT", "Notes"
    ]

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin",   color="CCCCCC"),
        right=Side(style="thin",  color="CCCCCC"),
        top=Side(style="thin",    color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_num, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_num, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = border

    sample_data = [
        ["2020-04-15", "BUY",     "TCS",      "INE467B01029", 100, 200000, 200, 0,   200, "Sample buy with STT"],
        ["2021-06-20", "BUY",     "RELIANCE", "INE002A01018",  50, 100000, 150, 0,   100, "Sample buy with STT"],
        ["2022-03-01", "SPLIT",   "RELIANCE", "",               1,      2,   0, 0,     0, "2:1 split — Amount=ratio, Qty=1"],
        ["2023-08-10", "BONUS",   "TCS",      "INE467B01029", 100,      0,   0, 0,     0, "1:1 Bonus issue — no STT"],
        ["2023-09-01", "GIFT",    "INFY",     "INE009A01021",  50,  75000,   0, 0,     0, "Gift from father — no STT"],
        ["2023-10-15", "INHERIT", "HDFCBANK", "INE040A01034",  30,  60000,   0, 0,     0, "Inherited — no STT"],
        ["2024-08-15", "SELL",    "TCS",      "INE467B01029",  50, 175000,   0, 150, 175, "Sample sell with STT"],
    ]

    for row_num, row_data in enumerate(sample_data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_num, column=col_num, value=value)
            cell.border    = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    type_validation = DataValidation(
        type="list",
        formula1='"BUY,SELL,BUYBACK,BONUS,SPLIT,GIFT,INHERIT"',
        allow_blank=True
    )
    type_validation.error       = "Please select a valid type"
    type_validation.errorTitle  = "Invalid Transaction Type"
    type_validation.prompt      = "Click arrow to select type"
    type_validation.promptTitle = "Transaction Type"

    ws.add_data_validation(type_validation)
    type_validation.add("B2:B1000")

    column_widths = [12, 12, 18, 16, 10, 14, 14, 14, 10, 35]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # ── Instructions sheet ─────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["CAPITAL GAINS TOOL — EXCEL IMPORT INSTRUCTIONS"],
        [""],
        ["1. Use the 'Transactions' sheet to enter your data."],
        ["2. DELETE the sample rows before adding your real data."],
        ["3. Date format: YYYY-MM-DD  (example: 2024-08-15)"],
        ["4. Type: Click the cell to see the dropdown."],
        ["   Options: BUY, SELL, BUYBACK, BONUS, SPLIT, GIFT, INHERIT"],
        ["5. Company name in CAPITAL letters (e.g. RELIANCE not Reliance)"],
        ["6. Amount = TOTAL value (qty × price per share)"],
        ["   Example: 100 shares at Rs 500 = Amount 50000"],
        ["7. Buy_Expenses = Brokerage + fees on BUY side (EXCLUDE STT)"],
        ["8. Sell_Expenses = Brokerage + fees on SELL side (EXCLUDE STT)"],
        ["9. STT = Securities Transaction Tax (separate column)"],
        ["   STT is INFORMATIONAL only — Section 40(a)(ib) says STT is NOT"],
        ["   deductible from capital gains. We track it for records only."],
        ["10. ISIN and Notes are optional"],
        [""],
        ["SPECIAL TYPES:"],
        [""],
        ["BONUS SHARES (Sec 55):"],
        ["   Date = Bonus allotment date"],
        ["   Amount = 0  (cost is zero)"],
        ["   STT = 0  (no transaction)"],
        ["   Grandfathering never applies to bonus shares"],
        [""],
        ["STOCK SPLIT:"],
        ["   Date = Date the split took effect"],
        ["   Amount = Split ratio as a decimal (e.g. 2 for 2:1, 5 for 5:1)"],
        ["   Quantity = 1 (not used — just put 1)"],
        ["   STT = 0  (no actual trade)"],
        ["   Effect: All buy lots before this date get qty × ratio"],
        [""],
        ["GIFT (Sec 49(1)):"],
        ["   Date = Donor's ORIGINAL purchase date (not the gift date)"],
        ["   Amount = Donor's ORIGINAL cost (what donor paid)"],
        ["   STT = 0  (no trade by you — was gift)"],
        ["   Use Notes to mention gift date and donor name"],
        [""],
        ["INHERIT (Sec 49(1)):"],
        ["   Date = Original owner's purchase date (deceased person's date)"],
        ["   Amount = Original owner's cost of acquisition"],
        ["   STT = 0  (no trade by you — was inherited)"],
        ["   Use Notes to mention inheritance date and relationship"],
        [""],
        ["STT QUICK GUIDE (for reference only):"],
        ["   Delivery Buy:  0.1% of transaction value"],
        ["   Delivery Sell: 0.1% of transaction value"],
        ["   Buyback:       0.2% of buyback value"],
        ["   Always check the broker contract note for exact STT paid"],
        [""],
        ["After filling: Save → Upload in the app"]
    ]

    for row_num, line in enumerate(instructions, start=1):
        cell = ws2.cell(row=row_num, column=1, value=line[0])
        if row_num == 1:
            cell.font = Font(bold=True, size=14, color="1A237E")
        elif "SPECIAL TYPES" in line[0] or "STT QUICK" in line[0]:
            cell.font = Font(bold=True, size=12, color="6A1B9A")
        elif line[0].startswith("BONUS") or line[0].startswith("STOCK") \
                or line[0].startswith("GIFT") or line[0].startswith("INHERIT"):
            cell.font = Font(bold=True, size=11, color="1B5E20")
        else:
            cell.font = Font(size=11)

    ws2.column_dimensions["A"].width = 80

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def read_excel_file(file_storage):
    """
    Read Excel file and return transactions.
    Returns: (transactions, errors, warnings)
      - errors   = rows that were SKIPPED (red)
      - warnings = rows that were IMPORTED but with issues to verify (yellow)
    """
    errors       = []
    warnings     = []
    transactions = []

    try:
        df = pd.read_excel(file_storage, sheet_name="Transactions")
    except ValueError:
        errors.append("Sheet 'Transactions' not found. Use the official template.")
        return [], errors, warnings
    except Exception as e:
        errors.append(f"Could not read the file: {str(e)}")
        return [], errors, warnings

    required_cols = ["Date", "Type", "Company", "Quantity", "Amount"]
    missing = [c for c in required_cols if c not in df.columns]

    if missing:
        errors.append(f"Missing columns: {', '.join(missing)}. Use the template.")
        return [], errors, warnings

    VALID_TYPES = ["BUY", "SELL", "BUYBACK", "BONUS", "SPLIT", "GIFT", "INHERIT"]

    for index, row in df.iterrows():
        row_num = index + 2

        if pd.isna(row["Date"]) and pd.isna(row["Company"]):
            continue

        try:
            if pd.isna(row["Date"]):
                errors.append(f"Row {row_num}: Date is missing")
                continue
            if pd.isna(row["Type"]):
                errors.append(f"Row {row_num}: Type is missing")
                continue
            if pd.isna(row["Company"]):
                errors.append(f"Row {row_num}: Company is missing")
                continue
            if pd.isna(row["Quantity"]):
                errors.append(f"Row {row_num}: Quantity is missing")
                continue
            if pd.isna(row["Amount"]):
                errors.append(f"Row {row_num}: Amount is missing")
                continue

            tx_type = str(row["Type"]).upper().strip()
            if tx_type not in VALID_TYPES:
                errors.append(
                    f"Row {row_num}: Type must be one of "
                    f"{', '.join(VALID_TYPES)} (found: '{tx_type}')"
                )
                continue

            date_value = row["Date"]
            if isinstance(date_value, pd.Timestamp):
                date_str = date_value.strftime("%Y-%m-%d")
            else:
                date_str = str(date_value).split(" ")[0]

            quantity = int(row["Quantity"])
            amount   = round(float(row["Amount"]), 2)

            if quantity <= 0:
                errors.append(f"Row {row_num}: Quantity must be > 0")
                continue

            if tx_type == "BONUS":
                amount = 0.0

            elif tx_type == "SPLIT":
                if amount <= 0:
                    errors.append(
                        f"Row {row_num}: SPLIT ratio (Amount column) "
                        f"must be > 0 (e.g. 2 for a 2:1 split)"
                    )
                    continue

            elif tx_type in ["GIFT", "INHERIT"]:
                if amount < 0:
                    errors.append(
                        f"Row {row_num}: Amount cannot be negative"
                    )
                    continue

            elif tx_type == "BUY":
                if amount < 0:
                    errors.append(
                        f"Row {row_num}: Amount cannot be negative"
                    )
                    continue
                if amount == 0:
                    warnings.append(
                        f"Row {row_num}: BUY has Amount = 0 — "
                        f"if this is a Bonus share, please change "
                        f"Type to BONUS on the review page"
                    )

            elif tx_type in ["SELL", "BUYBACK"]:
                if amount <= 0:
                    errors.append(
                        f"Row {row_num}: {tx_type} must have Amount > 0"
                    )
                    continue

            isin = (
                str(row.get("ISIN", ""))
                if not pd.isna(row.get("ISIN", "")) else ""
            )
            buy_expenses = round(
                float(row.get("Buy_Expenses", 0) or 0)
                if not pd.isna(row.get("Buy_Expenses", 0)) else 0,
                2
            )
            sell_expenses = round(
                float(row.get("Sell_Expenses", 0) or 0)
                if not pd.isna(row.get("Sell_Expenses", 0)) else 0,
                2
            )

            stt = 0.0
            if "STT" in df.columns:
                stt_value = row.get("STT", 0)
                if not pd.isna(stt_value):
                    try:
                        stt = round(float(stt_value or 0), 2)
                        if stt < 0:
                            warnings.append(
                                f"Row {row_num}: STT was negative "
                                f"(found: {stt}) — set to 0"
                            )
                            stt = 0
                    except (ValueError, TypeError):
                        stt = 0

            notes = (
                str(row.get("Notes", ""))
                if not pd.isna(row.get("Notes", "")) else ""
            )

            transactions.append({
                "date"         : date_str,
                "type"         : tx_type,
                "company"      : str(row["Company"]).upper().strip(),
                "isin"         : isin.strip(),
                "quantity"     : quantity,
                "amount"       : amount,
                "buy_expenses" : buy_expenses,
                "sell_expenses": sell_expenses,
                "stt"          : stt,
                "notes"        : notes.strip()
            })

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
            continue

    return transactions, errors, warnings


# ============================================================
# FMV TEMPLATE
# ============================================================

def generate_fmv_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "FMV_Data"

    headers = ["Company", "FMV_31_Jan_2018"]

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin",   color="CCCCCC"),
        right=Side(style="thin",  color="CCCCCC"),
        top=Side(style="thin",    color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_num, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_num, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = border

    sample_data = [
        ["RELIANCE",  953.10],
        ["TCS",      3068.55],
        ["INFY",     1162.50],
        ["HDFCBANK", 1888.30],
        ["WIPRO",     309.95],
    ]

    for row_num, row_data in enumerate(sample_data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_num, column=col_num, value=value)
            cell.border    = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["FMV BULK UPLOAD — INSTRUCTIONS"], [""],
        ["1. Use the 'FMV_Data' sheet to enter your data."],
        ["2. DELETE the sample rows before adding your real data."],
        ["3. Column A: Company name in CAPITAL LETTERS"],
        ["4. Column B: Closing price on 31-Jan-2018 (in Rupees)"],
        ["5. Use the HIGHER of BSE/NSE closing price."], [""],
        ["WHERE TO FIND FMV VALUES:"], [""],
        ["BSE: https://www.bseindia.com → Historical Data → 31-Jan-2018"],
        ["NSE: https://www.nseindia.com → Historical Data → 31-Jan-2018"], [""],
        ["If company already exists, FMV will be UPDATED."], [""],
        ["After filling: Save the file → Upload in the FMV card"]
    ]

    for row_num, line in enumerate(instructions, start=1):
        cell = ws2.cell(row=row_num, column=1, value=line[0])
        if row_num == 1:
            cell.font = Font(bold=True, size=14, color="6A1B9A")
        else:
            cell.font = Font(size=11)

    ws2.column_dimensions["A"].width = 80

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def read_fmv_file(file_storage):
    errors      = []
    fmv_entries = []

    try:
        df = pd.read_excel(file_storage, sheet_name="FMV_Data")
    except ValueError:
        errors.append("Sheet 'FMV_Data' not found. Use the FMV template.")
        return [], errors
    except Exception as e:
        errors.append(f"Could not read the file: {str(e)}")
        return [], errors

    required_cols = ["Company", "FMV_31_Jan_2018"]
    missing = [c for c in required_cols if c not in df.columns]

    if missing:
        errors.append(f"Missing columns: {', '.join(missing)}. Use the FMV template.")
        return [], errors

    for index, row in df.iterrows():
        row_num = index + 2

        if pd.isna(row["Company"]) and pd.isna(row["FMV_31_Jan_2018"]):
            continue

        try:
            if pd.isna(row["Company"]):
                errors.append(f"Row {row_num}: Company name is missing")
                continue
            if pd.isna(row["FMV_31_Jan_2018"]):
                errors.append(f"Row {row_num}: FMV value is missing")
                continue

            company = str(row["Company"]).upper().strip()
            fmv     = float(row["FMV_31_Jan_2018"])

            if not company:
                errors.append(f"Row {row_num}: Company name is empty")
                continue
            if fmv <= 0:
                errors.append(f"Row {row_num}: FMV must be > 0 (found: {fmv})")
                continue

            fmv_entries.append({"company": company, "fmv": fmv})

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
            continue

    return fmv_entries, errors


# ============================================================
# LAYOUT 1 — CONSOLIDATED PORTFOLIO EXCEL (Combined Date)
# One Buy Date column + each client gets Qty + Rate
# ============================================================
# CLEAN B&W LOOK — no fill colors, only borders + bold headers
# ============================================================

def generate_consolidated_excel(client_names, lots, client_ids):
    """
    Generate a lot-wise consolidated portfolio Excel file (Layout 1).
    Layout: Company | ISIN | Buy Date | C1 Qty | C1 Rate | C2 Qty | C2 Rate | ...

    Args:
        client_names : dict { client_id -> client_name }
        lots         : list of dicts (each = one buy lot row)
        client_ids   : list of client IDs in display order
    Returns:
        BytesIO ready to send as file download
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Portfolio"

    # ── Styles (B&W only) ──────────────────────────────────────
    border = Border(
        left=Side(style="thin",   color="000000"),
        right=Side(style="thin",  color="000000"),
        top=Side(style="thin",    color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center")
    right       = Alignment(horizontal="right",  vertical="center")
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    bold_font   = Font(bold=True, size=10)

    # ── Header Rows 1 + 2 ──────────────────────────────────────
    fixed_headers = ["Company", "ISIN", "Buy Date"]

    for col_num, header in enumerate(fixed_headers, start=1):
        top_cell           = ws.cell(row=1, column=col_num, value=header)
        top_cell.font      = header_font
        top_cell.alignment = center
        top_cell.border    = border

        bottom_cell           = ws.cell(row=2, column=col_num)
        bottom_cell.font      = header_font
        bottom_cell.alignment = center
        bottom_cell.border    = border

        ws.merge_cells(
            start_row=1, start_column=col_num,
            end_row=2,   end_column=col_num
        )

    # Client headers — Qty + Rate per client
    col = 4
    for cid in client_ids:
        name = client_names.get(cid, f"Client {cid}")

        left_cell           = ws.cell(row=1, column=col, value=name)
        left_cell.font      = header_font
        left_cell.alignment = center
        left_cell.border    = border

        right_cell           = ws.cell(row=1, column=col + 1)
        right_cell.font      = header_font
        right_cell.alignment = center
        right_cell.border    = border

        ws.merge_cells(
            start_row=1, start_column=col,
            end_row=1,   end_column=col + 1
        )

        # Row 2: Qty | Rate (₹)
        for label, offset in [("Qty", 0), ("Rate (₹)", 1)]:
            cc            = ws.cell(row=2, column=col + offset, value=label)
            cc.font       = header_font
            cc.alignment  = center
            cc.border     = border

        col += 2

    # ── Data Rows ──────────────────────────────────────────────
    data_row     = 3
    last_company = None

    for lot in lots:
        company     = lot["company"]
        isin        = lot["isin"]
        buy_date    = lot["buy_date"]
        client_lots = lot["client_lots"]

        is_new_company = (company != last_company)
        if is_new_company:
            last_company = company

        c1            = ws.cell(row=data_row, column=1, value=company)
        c1.alignment  = left
        c1.border     = border
        c1.font       = bold_font if is_new_company else normal_font

        c2            = ws.cell(row=data_row, column=2,
                                value=isin if isin else "")
        c2.alignment  = center
        c2.border     = border
        c2.font       = normal_font

        c3            = ws.cell(row=data_row, column=3, value=buy_date)
        c3.alignment  = center
        c3.border     = border
        c3.font       = normal_font

        col = 4
        for cid in client_ids:
            if cid in client_lots:
                qty  = client_lots[cid]["qty"]
                rate = client_lots[cid]["rate"]
            else:
                qty  = ""
                rate = ""

            qc            = ws.cell(row=data_row, column=col, value=qty)
            qc.alignment  = right
            qc.border     = border
            qc.font       = normal_font

            rc            = ws.cell(row=data_row, column=col + 1, value=rate)
            rc.alignment  = right
            rc.border     = border
            rc.font       = normal_font

            if isinstance(rate, float):
                rc.number_format = '#,##0.00'

            col += 2

        data_row += 1

    # ── Column Widths ──────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 13

    col_idx = 4
    for _ in client_ids:
        ws.column_dimensions[_col_letter(col_idx)].width     = 10
        ws.column_dimensions[_col_letter(col_idx + 1)].width = 13
        col_idx += 2

    ws.freeze_panes = "D3"

    # ── Footer Note ────────────────────────────────────────────
    footer_row = data_row + 1
    note = ws.cell(
        row=footer_row, column=1,
        value=(
            "Generated by Capital Gains AI Tool  |  "
            "Holdings as of today  |  Each row = one buy lot  |  "
            "Rate = (Cost + Buy Expenses) / Qty  |  FIFO method applied"
        )
    )
    note.font      = Font(italic=True, size=9)
    note.alignment = left

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# LAYOUT 2 — CONSOLIDATED PORTFOLIO EXCEL (Per-Client Date)
# ============================================================
# FIXED (Day 12.6):
#   - Company name + ISIN repeat on EVERY row
#   - Same company + same date across clients = SAME row
#   - Different dates = different rows
# CLEAN B&W LOOK — no fill colors, only borders + bold headers
# ============================================================

def generate_consolidated_excel_perclient_dates(client_names, client_ids,
                                                 client_lots_map):
    """
    Generate Layout 2 — each client gets their own (Qty + Date + Rate) cols.
    Same company + same date across clients = SAME row.

    Args:
        client_names    : dict { client_id -> client_name }
        client_ids      : list of client IDs in display order
        client_lots_map : dict { client_id -> list of lots }
                          Each lot has: company, isin, buy_date, buy_date_raw,
                          quantity, rate_per_share

    Returns:
        BytesIO ready for download
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Per-Client Holdings"

    # ── Styles (B&W only) ──────────────────────────────────────
    border = Border(
        left=Side(style="thin",   color="000000"),
        right=Side(style="thin",  color="000000"),
        top=Side(style="thin",    color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center")
    right       = Alignment(horizontal="right",  vertical="center")
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    bold_font   = Font(bold=True, size=10)

    # ── Step 1: Build data structure ───────────────────────────
    company_data = {}

    for cid in client_ids:
        lots_for_client = client_lots_map.get(cid, [])

        for lot in lots_for_client:
            company       = lot["company"]
            isin          = lot["isin"]
            buy_date      = lot["buy_date"]
            buy_date_raw  = lot["buy_date_raw"]
            qty           = lot["quantity"]
            rate          = lot["rate_per_share"]

            # Use ISIN as key if available, else company name
            key = isin if isin else company

            if key not in company_data:
                company_data[key] = {
                    "company": company,
                    "isin"   : isin,
                    "rows"   : {},
                }

            if buy_date_raw not in company_data[key]["rows"]:
                company_data[key]["rows"][buy_date_raw] = {
                    "display_date": buy_date,
                    "clients"     : {},
                }

            # If the same client somehow has multiple lots on same date for
            # same company (rare), accumulate quantities + weighted avg rate
            if cid in company_data[key]["rows"][buy_date_raw]["clients"]:
                existing = company_data[key]["rows"][buy_date_raw]["clients"][cid]
                total_qty  = existing["qty"] + qty
                total_cost = (existing["qty"] * existing["rate"]) + (qty * rate)
                avg_rate   = total_cost / total_qty if total_qty > 0 else 0
                company_data[key]["rows"][buy_date_raw]["clients"][cid] = {
                    "qty" : total_qty,
                    "rate": avg_rate,
                }
            else:
                company_data[key]["rows"][buy_date_raw]["clients"][cid] = {
                    "qty" : qty,
                    "rate": rate,
                }

    # ── Step 2: Build headers ──────────────────────────────────
    fixed_headers = ["Company", "ISIN"]
    for col_num, header in enumerate(fixed_headers, start=1):
        top_cell           = ws.cell(row=1, column=col_num, value=header)
        top_cell.font      = header_font
        top_cell.alignment = center
        top_cell.border    = border

        bottom_cell           = ws.cell(row=2, column=col_num)
        bottom_cell.font      = header_font
        bottom_cell.alignment = center
        bottom_cell.border    = border

        ws.merge_cells(
            start_row=1, start_column=col_num,
            end_row=2,   end_column=col_num
        )

    # Client headers — 3 columns each
    col = 3
    for cid in client_ids:
        name = client_names.get(cid, f"Client {cid}")

        # Style all 3 cells BEFORE merging
        for c in [col, col + 1, col + 2]:
            cc           = ws.cell(row=1, column=c)
            cc.font      = header_font
            cc.alignment = center
            cc.border    = border

        ws.cell(row=1, column=col).value = name

        ws.merge_cells(
            start_row=1, start_column=col,
            end_row=1,   end_column=col + 2
        )

        # Row 2: Qty | Buy Date | Rate (₹)
        for label, offset in [("Qty", 0), ("Buy Date", 1), ("Rate (₹)", 2)]:
            cc            = ws.cell(row=2, column=col + offset, value=label)
            cc.font       = header_font
            cc.alignment  = center
            cc.border     = border

        col += 3

    # ── Step 3: Sort companies alphabetically ──────────────────
    sorted_company_keys = sorted(
        company_data.keys(),
        key=lambda k: company_data[k]["company"]
    )

    # ── Step 4: Write data rows ────────────────────────────────
    data_row = 3

    for comp_key in sorted_company_keys:
        comp_info    = company_data[comp_key]
        company_name = comp_info["company"]
        isin         = comp_info["isin"]
        rows_by_date = comp_info["rows"]

        # Sort dates oldest first
        sorted_dates = sorted(rows_by_date.keys())

        for date_raw in sorted_dates:
            row_info        = rows_by_date[date_raw]
            display_date    = row_info["display_date"]
            clients_on_date = row_info["clients"]

            # Company name — on EVERY row
            c1            = ws.cell(row=data_row, column=1, value=company_name)
            c1.alignment  = left
            c1.border     = border
            c1.font       = bold_font

            # ISIN — on EVERY row
            c2            = ws.cell(row=data_row, column=2,
                                    value=isin if isin else "")
            c2.alignment  = center
            c2.border     = border
            c2.font       = normal_font

            # Each client's columns
            col = 3
            for cid in client_ids:
                if cid in clients_on_date:
                    qty      = clients_on_date[cid]["qty"]
                    rate     = round(clients_on_date[cid]["rate"], 2)
                    date_val = display_date
                else:
                    qty      = ""
                    rate     = ""
                    date_val = ""

                # Qty
                qc            = ws.cell(row=data_row, column=col, value=qty)
                qc.alignment  = right
                qc.border     = border
                qc.font       = normal_font

                # Buy Date
                dc            = ws.cell(row=data_row, column=col + 1,
                                        value=date_val)
                dc.alignment  = center
                dc.border     = border
                dc.font       = normal_font

                # Rate
                rc            = ws.cell(row=data_row, column=col + 2,
                                        value=rate)
                rc.alignment  = right
                rc.border     = border
                rc.font       = normal_font

                if isinstance(rate, float):
                    rc.number_format = '#,##0.00'

                col += 3

            data_row += 1

    # ── Column Widths ──────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16

    col_idx = 3
    for _ in client_ids:
        ws.column_dimensions[_col_letter(col_idx)].width     = 9    # Qty
        ws.column_dimensions[_col_letter(col_idx + 1)].width = 13   # Date
        ws.column_dimensions[_col_letter(col_idx + 2)].width = 13   # Rate
        col_idx += 3

    ws.freeze_panes = "C3"

    # ── Footer Note ────────────────────────────────────────────
    footer_row = data_row + 1
    note = ws.cell(
        row=footer_row, column=1,
        value=(
            "Generated by Capital Gains AI Tool  |  "
            "Holdings as of today  |  Each row = one date for one company  |  "
            "Same date across clients = same row  |  "
            "Rate = (Cost + Buy Expenses) / Qty  |  FIFO method applied"
        )
    )
    note.font      = Font(italic=True, size=9)
    note.alignment = left

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _col_letter(col_index):
    """
    Convert a column number to Excel column letter(s).
    1=A, 2=B, 26=Z, 27=AA, 28=AB, etc.
    """
    result = ""
    while col_index > 0:
        col_index, remainder = divmod(col_index - 1, 26)
        result = chr(65 + remainder) + result
    return result