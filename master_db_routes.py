# master_db_routes.py
# ============================================================
# Day 16 — Master Securities Database routes
# Phase 2: CSV upload + browse + manual add/edit/delete
# Phase 3: Match API for PDF/Excel review page
# ============================================================

from flask import (
    Blueprint, render_template, request,
    redirect, url_for, flash, jsonify
)
import csv
import io

from database import (
    add_security,
    get_all_securities,
    get_security_by_isin,
    delete_security,
    add_alias_to_security,
    check_alias_conflict,
    search_securities_by_name,
)

master_db_bp = Blueprint("master_db", __name__)


# ============================================================
# MAIN PAGE — Browse + Search
# ============================================================
@master_db_bp.route("/master_db")
def master_db_page():
    search   = request.args.get("search", "").strip()
    page     = int(request.args.get("page", 1) or 1)
    per_page = 50

    securities, total_count = get_all_securities(
        search=search if search else None,
        page=page,
        per_page=per_page
    )

    total_pages = (total_count + per_page - 1) // per_page

    return render_template(
        "master_db.html",
        securities  = securities,
        total_count = total_count,
        search      = search,
        page        = page,
        total_pages = total_pages,
    )


# ============================================================
# UPLOAD CSV (NSE or BSE)
# ============================================================
@master_db_bp.route("/master_db/upload_csv", methods=["POST"])
def upload_csv():
    if "csv_file" not in request.files:
        flash("❌ No file uploaded.", "error")
        return redirect(url_for("master_db.master_db_page"))

    file = request.files["csv_file"]

    if not file or file.filename == "":
        flash("❌ No file selected.", "error")
        return redirect(url_for("master_db.master_db_page"))

    if not file.filename.lower().endswith((".csv", ".xlsx", ".xls")):
        flash("❌ Only CSV or Excel files allowed.", "error")
        return redirect(url_for("master_db.master_db_page"))

    try:
        rows = _read_file_to_rows(file)
    except Exception as e:
        flash(f"❌ Could not read file: {str(e)}", "error")
        return redirect(url_for("master_db.master_db_page"))

    if not rows:
        flash("❌ File appears to be empty.", "error")
        return redirect(url_for("master_db.master_db_page"))

    # Detect format — safely handle None keys
    headers = []
    for h in rows[0].keys():
        if h is not None:
            headers.append(str(h).strip().upper())

    format_type = _detect_format(headers)

    if not format_type:
        flash(
            "❌ Unrecognized file format. "
            "Expected NSE columns (SYMBOL, NAME OF COMPANY, ISIN NUMBER) "
            "or BSE columns (Security Id, Security Name, ISIN No). "
            f"Got: {', '.join(headers[:8])}",
            "error"
        )
        return redirect(url_for("master_db.master_db_page"))

    imported = 0
    skipped  = 0
    errors   = 0

    for row in rows:
        try:
            isin, name, alias, exchange = _parse_row(row, format_type)

            if not isin or len(isin) != 12:
                errors += 1
                continue

            if not name:
                errors += 1
                continue

            existing = get_security_by_isin(isin)

            if existing:
                if alias and alias not in existing["aliases"]:
                    add_alias_to_security(isin, alias)
                skipped += 1
                continue

            aliases = [alias] if alias else []
            add_security(isin, name, aliases, exchange)
            imported += 1

        except Exception:
            errors += 1
            continue

    msg = (
        f"✅ Imported {imported} new securities. "
        f"Skipped {skipped} duplicates."
    )
    if errors:
        msg += f" ⚠️ {errors} rows had errors and were skipped."

    flash(msg, "success")
    return redirect(url_for("master_db.master_db_page"))


# ============================================================
# MANUAL ADD
# ============================================================
@master_db_bp.route("/master_db/add", methods=["POST"])
def manual_add():
    isin     = request.form.get("isin", "").strip().upper()
    name     = request.form.get("official_name", "").strip().upper()
    exchange = request.form.get("exchange", "").strip().upper() or None
    aliases_raw = request.form.get("aliases", "").strip()

    if not isin or len(isin) != 12:
        flash("❌ ISIN must be exactly 12 characters.", "error")
        return redirect(url_for("master_db.master_db_page"))

    if not name:
        flash("❌ Official name required.", "error")
        return redirect(url_for("master_db.master_db_page"))

    existing = get_security_by_isin(isin)
    if existing:
        flash(
            f"⚠️ ISIN already exists: {existing['official_name']}. "
            f"Use Edit instead.",
            "error"
        )
        return redirect(url_for("master_db.master_db_page"))

    aliases = []
    if aliases_raw:
        aliases = [a.strip().upper() for a in aliases_raw.split(",") if a.strip()]

    for alias in aliases:
        conflict_isin = check_alias_conflict(alias)
        if conflict_isin:
            conflict_sec = get_security_by_isin(conflict_isin)
            flash(
                f"❌ Alias '{alias}' already used by "
                f"{conflict_sec['official_name']} ({conflict_isin}). "
                f"Cannot add.",
                "error"
            )
            return redirect(url_for("master_db.master_db_page"))

    add_security(isin, name, aliases, exchange)
    flash(f"✅ Added {name} to master DB.", "success")
    return redirect(url_for("master_db.master_db_page"))


# ============================================================
# EDIT
# ============================================================
@master_db_bp.route("/master_db/edit/<isin>", methods=["POST"])
def edit_security(isin):
    isin = isin.upper().strip()

    existing = get_security_by_isin(isin)
    if not existing:
        flash("❌ Security not found.", "error")
        return redirect(url_for("master_db.master_db_page"))

    name        = request.form.get("official_name", "").strip().upper()
    exchange    = request.form.get("exchange", "").strip().upper() or None
    aliases_raw = request.form.get("aliases", "").strip()

    if not name:
        flash("❌ Official name required.", "error")
        return redirect(url_for("master_db.master_db_page"))

    aliases = []
    if aliases_raw:
        aliases = [a.strip().upper() for a in aliases_raw.split(",") if a.strip()]

    for alias in aliases:
        conflict_isin = check_alias_conflict(alias)
        if conflict_isin and conflict_isin != isin:
            conflict_sec = get_security_by_isin(conflict_isin)
            flash(
                f"❌ Alias '{alias}' already used by "
                f"{conflict_sec['official_name']} ({conflict_isin}).",
                "error"
            )
            return redirect(url_for("master_db.master_db_page"))

    add_security(isin, name, aliases, exchange)
    flash(f"✅ Updated {name}.", "success")
    return redirect(url_for("master_db.master_db_page"))


# ============================================================
# DELETE
# ============================================================
@master_db_bp.route("/master_db/delete/<isin>", methods=["POST"])
def delete_security_route(isin):
    isin = isin.upper().strip()

    existing = get_security_by_isin(isin)
    if not existing:
        flash("❌ Security not found.", "error")
        return redirect(url_for("master_db.master_db_page"))

    delete_security(isin)
    flash(f"🗑️ Deleted {existing['official_name']}.", "success")
    return redirect(url_for("master_db.master_db_page"))


# ============================================================
# PHASE 3 — API: Match company name against master DB
# Returns JSON list of suggestions for PDF/Excel review page
# ============================================================
@master_db_bp.route("/api/master_match")
def api_master_match():
    """
    Query params:
      name = company name to match
      isin = optional ISIN (if present, ISIN match wins)

    Returns JSON:
      {
        "matches": [
          {"isin": "...", "official_name": "...", "match_type": "...", "exchange": "..."},
          ...
        ]
      }
    """
    name = (request.args.get("name") or "").strip()
    isin = (request.args.get("isin") or "").strip().upper()

    matches = []

    # Priority 1: ISIN match (most trustworthy)
    if isin and len(isin) == 12:
        sec = get_security_by_isin(isin)
        if sec:
            matches.append({
                "isin"         : sec["isin"],
                "official_name": sec["official_name"],
                "match_type"   : "isin_exact",
                "exchange"     : sec["exchange"] or ""
            })
            return jsonify({"matches": matches})

    # Priority 2: Name-based match
    if name:
        name_matches = search_securities_by_name(name)
        # De-duplicate by ISIN (in case both alias and name match)
        seen = set()
        for m in name_matches:
            if m["isin"] not in seen:
                seen.add(m["isin"])
                matches.append({
                    "isin"         : m["isin"],
                    "official_name": m["official_name"],
                    "match_type"   : m["match_type"],
                    "exchange"     : m.get("exchange") or ""
                })

    return jsonify({"matches": matches})


# ============================================================
# HELPERS
# ============================================================

def _read_file_to_rows(file):
    """Read CSV or Excel into list of dicts. Safely handles empty cells."""
    filename = file.filename.lower()

    if filename.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        rows_data = []
        headers = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                for c in row:
                    if c is not None and str(c).strip():
                        headers.append(str(c).strip())
                    else:
                        headers.append("")
            else:
                row_dict = {}
                for h, v in zip(headers, row):
                    if h:
                        row_dict[h] = str(v).strip() if v is not None else ""
                if row_dict:
                    rows_data.append(row_dict)

        return rows_data

    else:
        content = file.read()
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")

        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            clean_row = {}
            for k, v in row.items():
                if k is not None and str(k).strip():
                    clean_row[str(k).strip()] = (v or "").strip()
            if clean_row:
                rows.append(clean_row)
        return rows


def _detect_format(headers):
    headers_upper = [h.upper() for h in headers if h]
    headers_set = set(headers_upper)

    if "ISIN NUMBER" in headers_set and "NAME OF COMPANY" in headers_set:
        return "NSE"

    if "ISIN NO" in headers_set and (
        "SECURITY NAME" in headers_set or "ISSUER NAME" in headers_set
    ):
        return "BSE"

    return None


def _parse_row(row, format_type):
    row_upper = {}
    for k, v in row.items():
        if k is not None:
            key_clean = str(k).upper().strip()
            row_upper[key_clean] = str(v or "").strip()

    if format_type == "NSE":
        isin     = row_upper.get("ISIN NUMBER", "").upper()
        name     = row_upper.get("NAME OF COMPANY", "").upper()
        alias    = row_upper.get("SYMBOL", "").upper()
        exchange = "NSE"

    elif format_type == "BSE":
        isin     = row_upper.get("ISIN NO", "").upper()
        name     = (
            row_upper.get("ISSUER NAME", "")
            or row_upper.get("SECURITY NAME", "")
        ).upper()
        alias    = row_upper.get("SECURITY ID", "").upper()
        exchange = "BSE"

    else:
        return None, None, None, None

    return isin, name, alias, exchange